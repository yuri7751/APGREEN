import queue
import math
import time
import serial
from pymodbus.client import ModbusSerialClient
import dearpygui.dearpygui as dpg
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import os
import threading
import struct
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="matplotlib")
import pandas as pd
import json
import os
import multiprocessing
SETTINGS_FILE = "settings.json"
default_config = {
    "RS485_PORT": "COM6",
    "GAS_ANALYZER_PORT": "COM4",
    "PM_PORT": "COM3"
    # Add other default settings as needed
}

if not os.path.exists(SETTINGS_FILE):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(default_config, f, indent=2)

with open(SETTINGS_FILE) as f:
    config = json.load(f)
    
SERIAL_PORT = 'COM6'
BAUDRATE = 9600
BYTESIZE = 8
TIMEOUT = 1
TK4_ADDRESSES = [1, 2, 3, 4]
TK4_RO_ADDRESSES = [5, 6]
COLOR_RED = [220, 50, 50]
COLOR_GREEN = [50, 180, 50]
COLOR_BLUE = [50, 120, 220]
COLOR_YELLOW = [220, 220, 0]
COLOR_WHITE = [255, 255, 255]
COLOR_GRAY = [200, 200, 200]
log_filename = f"process_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
with open("settings.json") as f:
    config = json.load(f)
SERIAL_PORT = config.get("RS485_PORT", "COM6")
GAS_ANALYZER_PORT = config.get("GAS_ANALYZER_PORT", "COM4")
PM_PORT = config.get("PM_PORT", "COM3")
class Logger:
    def __init__(self, filename):
        
        self.max_temp = 400.0
        self.max_press = 5.0
        self.filename = filename
        self.columns = [
            "Timestamp",
            "Heater1", "Heater2", "Heater3","Heater4'",
            "Temp1", "Temp2",
            "PSM4_1", "PSM4_2", "PSM4_3", "PSM4_4",
            "Power", "Energy",
            "MFM_Flow",
            "MFC1_Flow", "MFC2_Flow", "MFC3_Flow", "MFC4_Flow",
            "CO", "CO2", "CH4", "CnHm", "H2", "O2", "C2H2", "C2H4", "HHV", "N2",
        ]
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(self.columns)
            for i, col in enumerate(self.columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = 12
            wb.save(self.filename)

    def log(self, data, gas_values=None):
        row = [
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            *[(t if t is not None and t < 2000 else ("NC" if t == 31000 else t)) for t in (data.main_temps if hasattr(data, "main_temps") else [None]*4)],
            *[(t if t is not None and t < 2000 else ("NC" if t == 31000 else t)) for t in (data.ro_temps if hasattr(data, "ro_temps") else [None]*2)],
            *(data.pressures if hasattr(data, "pressures") else [None]*4),
            data.power if hasattr(data, "power") else None,
            data.energy if hasattr(data, "energy") else None,
            data.flow if hasattr(data, "flow") else None,
            *(data.mfc_flows if hasattr(data, "mfc_flows") else [None]*4),
            *(gas_values.get(gas) if gas_values else None for gas in ['CO', 'CO2', 'CH4', 'CnHm', 'H2', 'O2', 'C2H2', 'C2H4', 'HHV', 'N2'])
        ]
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            ws.append(row)
            wb.save(self.filename)
        except Exception as e:
            print(f"[Logger] Logging error: {e}")



# Fixed hardcoded ports
#TK4_PORT = 'COM5'
#PM_PORT = 'COM3'


# Colors - matching guiex.py exactly
COLOR_RED = [220, 50, 50]
COLOR_GREEN = [50, 180, 50]
COLOR_BLUE = [50, 120, 220]
COLOR_YELLOW = [220, 220, 0]
COLOR_WHITE = [255, 255, 255]
COLOR_GRAY = [200, 200, 200]

class AbnormalEventLogger:
    def __init__(self, filename="abnormal_events.log"):
        self.filename = filename
        # Create the file with headers if it doesn't exist
        if not os.path.exists(self.filename):
            with open(self.filename, 'w') as f:
                f.write("Timestamp,Event Description\n")

    def log_event(self, description):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(self.filename, 'a') as f:
            f.write(f"{timestamp},{description}\n")
        print(f"Logged event: {timestamp} - {description}")
class MFCController:
    def __init__(self, port=SERIAL_PORT, channels=[1,2,3,4]):
        self.port = port
        self.channels = channels

    def _checksum(self, data):
        cs = 0
        for b in data.encode():
            cs ^= b
        return f"{cs:02X}"

    def send_command(self, channel, cmd, addr, value=""):
        frame_wo_cs = f":{channel:02d}{cmd}{addr}{value}"
        cs = self._checksum(frame_wo_cs)
        frame = f"{frame_wo_cs}{cs}\r".encode()
        
        with serial.Serial(
            port=self.port, baudrate=BAUDRATE, bytesize=BYTESIZE,
            parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout=0.1
        ) as ser:
            ser.write(frame)
            resp = ser.read_until(b'\r')
            return resp

    def read_flow(self, channel):
        resp = self.send_command(channel, "03", "0038")
        if not resp or not resp.startswith(b':'):
            return None
        data_ascii = resp[7:15].decode()
        return struct.unpack('>f', bytes.fromhex(data_ascii))[0]

    def set_flow(self, channel, value):
    # Build IEEE754 float as 8 ASCII hex chars
        hexval = struct.pack('!f', float(value)).hex().upper()
    # Correct command for TSM-D: Command "01", DataType "07"
        resp = self.send_command(channel, "01", "07", hexval)
    # Optionally, check response for success:
        return resp.startswith(f":{channel:02d}81".encode())




    def on_off(self, channel, state):
        val = "01" if state else "00"
        resp = self.send_command(channel, "58", "02", val)
    # Response should start with :{channel}D8 for success
        return resp.startswith(f":{channel:02d}D8".encode())


    def read_all_flows(self):
        return [self.read_flow(ch) for ch in self.channels]

class PowerMeter:
    def __init__(self, port=PM_PORT):
        self.port = port
        self.ser = None
        self.lock = threading.Lock()
        
    def connect(self):
        """Connect and initialize power meter using correct commands"""
        try:
            self.ser = serial.Serial(
                port=self.port,
                baudrate=9600,
                bytesize=8,
                parity='N',
                stopbits=2,
                timeout=2,
                write_timeout=5,
                inter_byte_timeout=0.5
            )
            self._send(":COMMunicate:REMote ON")
            self._send(":NUMERIC:NORMAL:ITEM4 WH,1")
            self._send(":SCALING:CT:ELEMENT1 10")
            time.sleep(0.5)
            return True
        except Exception as e:
            print(f"PM connection failed: {str(e)}")
            if self.ser and self.ser.is_open:
                self.close()
            return False
    
    def _send(self, cmd, read_response=True):
        """Send command and get response"""
        with self.lock:
            self.ser.write(f"{cmd}\r\n".encode())
            time.sleep(0.2)
            if read_response:
                return self.ser.read_all().decode().strip()
            return None
    
    def is_connected(self):
        """Check if power meter is connected"""
        return self.ser is not None and self.ser.is_open
    
    def read_power(self):
        """Read active power using correct command"""
        try:
            response = self._send(":NUMERIC:NORMAL:VALUE?3")
            if response:
                return float(response)
            return 0.0
        except Exception as e:
            print(f"Power read error: {str(e)}")
            return 0.0
    
    def read_energy(self):
        """Read accumulated energy"""
        try:
            response = self._send(":NUMERIC:NORMAL:VALUE?4")
            if response:
                return float(response)
            return 0.0
        except Exception as e:
            print(f"Energy read error: {str(e)}")
            return 0.0
    
    def start_integration(self):
        """Start energy integration"""
        self.reset_integration()
        self._send(":INTEGrate:STARt", False)
    
    def stop_integration(self):
        """Stop energy integration"""
        self._send(":INTEGrate:STOP", False)
    
    def reset_integration(self):
        """Reset integration counters"""
        self._send(":INTEGrate:RESet", False)
        time.sleep(0.5)
    
    def close(self):
        """Close connection properly"""
        if self.ser and self.ser.is_open:
            self._send(":COMMunicate:REMote OFF", False)
            self.ser.close()
            self.ser = None

class TK4Controller:
    def __init__(self, client):
        self.client = client
        self.lock = threading.Lock()

    

    def close(self):
        if self.client:
            self.client.close()


            
    def read_temperature(self, address):
        """Read temperature from controller"""
        try:
            with self.lock:
                response = self.client.read_input_registers(
                    address=0x03E8, 
                    count=2, 
                    slave=address
                )
                if response.isError():
                    return None
                raw_pv, decimal_point = response.registers
                return raw_pv / (10 ** decimal_point)
        except Exception as e:
            print(f"Temperature read error: {str(e)}")
            return None
    
    def set_setpoint(self, address, temperature):
        """Set temperature setpoint"""
        try:
            with self.lock:
                response = self.client.write_register(
                    address=0x0000,
                    value=int(temperature ),  # 0.1°C resolution
                    slave=address
                )
                return not response.isError()
        except Exception as e:
            print(f"Set temperature error: {str(e)}")
            return False
    
    def control_heater(self, address, state):
        """Control heater state (on/off)"""
        try:
            with self.lock:
                response = self.client.write_register(
                    address=0x0032,
                    value=int(state),
                    slave=address
                )
                if response.isError():
                    print(f"Control heater error: {response}")
                    return False
                return True
        except Exception as e:
            print(f"Control heater error: {str(e)}")
        # Do NOT close the client here!
            return False

    
    def close(self):
        """Close connection"""
        if self.client:
            self.client.close()

class ModbusRelayController:
    def __init__(self, client, slave_id=8):
        self.client = client
        self.slave_id = slave_id
        self.lock = threading.Lock()

    def send_pulse(self, channel, duration=1.0):
        # This method is now synchronous and must be called only from the worker thread!
        try:
            self._write_channel(channel, 'on')
            time.sleep(duration)
            self._write_channel(channel, 'off')
        except Exception as e:
            print(f"Relay pulse error on channel {channel}: {e}")

    def _write_channel(self, channel, command):
        value = {'on': 0x0100, 'off': 0x0200}[command]
        print(f"Relay {channel} -> {command.upper()}")
        self.client.write_register(channel, value, slave=self.slave_id)

    def open_all(self):
        """Open (activate) all relay channels using the group command."""
        try:
        # 0x0000 register, value 0x0700 (per manual: Open all)
            self.client.write_register(0x0000, 0x0700, slave=self.slave_id)
            print("All relays OPEN (ON)")
        except Exception as e:
            print(f"Relay open_all error: {e}")

    def close_all(self):
        """Close (deactivate) all relay channels using the group command."""
        try:
        # 0x0000 register, value 0x0800 (per manual: Close all)
            self.client.write_register(0x0000, 0x0800, slave=self.slave_id)
            print("All relays CLOSED (OFF)")
        except Exception as e:
            print(f"Relay close_all error: {e}")


class PSM4Controller:
    def __init__(self, client):
        self.client = client  # Share TK4's Modbus client
        self.lock = threading.Lock()

    def read_pressures(self):
        """Read all 4 pressure channels (ID=7)"""
        pressures = []
        base_addrs = [0x03E8, 0x03ED, 0x03F2, 0x03F7]
        try:
            with self.lock:
                for addr in base_addrs:
                    response = self.client.read_input_registers(
                        address=addr + 1,
                        count=2,
                        slave=7
                    )
                    if response.isError():
                        pressures.append("NC")
                        continue
                    raw, decimal = response.registers
                # Correct scaling: divide by 10
                    val = raw / (10 ** decimal) / 10.0
                # If value > 20 bar, show NC
                    if val > 20.0:
                        pressures.append("NC")
                    else:
                        pressures.append(round(val, 2))

            return pressures
        except Exception as e:
            print(f"PSM4 read error: {str(e)}")
            return ["NC"]*4
    def close(self):
        if self.client:
            self.client.close()



# --- MFM (TSM-D) ---
import time
import serial
import struct

class MFMFlowMeter:
    def __init__(self, port='COM6', device_id=1):
        self.port = port
        self.device_id = device_id

    def _checksum(self, frame_wo_cs):
        cs = 0
        for b in frame_wo_cs.encode():
            cs ^= b
        return f"{cs:02X}"

    def build_read_flow_frame(self):
        id_ascii = f"{self.device_id:02X}"
        frame_wo_cs = f":{id_ascii}0300"
        cs = self._checksum(frame_wo_cs)
        return f"{frame_wo_cs}{cs}\r".encode()

    def read_flow(self):
        
        # Always open and close the port for each read, just like mfm_mfc1.py
        with serial.Serial(
            port=self.port, baudrate=BAUDRATE, bytesize=BYTESIZE,
            parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout=TIMEOUT
        ) as ser:
            frame = self.build_read_flow_frame()
                #print(f"[MFM] Sending: {frame!r}")
            ser.write(frame)
            resp = ser.read_until(b'\r')
                #print(f"[MFM] Raw response: {resp!r}")
            if not resp.startswith(b':'):
                raise ValueError("Invalid response")
            if len(resp) < 16:
                raise ValueError(f"Response too short: {len(resp)} bytes")
            try:
                data_ascii = resp[7:15].decode('ascii')
                data_bytes = bytes.fromhex(data_ascii)
                return struct.unpack('>f', data_bytes)[0]
            except Exception as e:
                print(f"[MFM] Hex dump: {resp.hex()}")
                raise e

class GasAnalyzer:
    def __init__(self, port="COM10"):
        self.port = port
        self.ser = None

    def connect(self):
        try:
            self.ser = serial.Serial(self.port, baudrate=9600, timeout=1)
            return True
        except Exception as e:
            print(f"Gas analyzer connection error: {e}")
            self.ser = None
            return False

    def close(self):
        if self.ser:
            self.ser.close()
            self.ser = None

    def read_gases(self):
        if not self.ser or not self.ser.is_open:
            if not self.connect():
                return None
        try:
            self.ser.reset_input_buffer()
            self.ser.write(bytes([0x11, 0x01, 0x01, 0xed]))
            time.sleep(0.2)
            nbytes = self.ser.in_waiting
            if nbytes >= 20:
                data = self.ser.read(nbytes)
                #print(f"Raw gas analyzer response: {data.hex()} ({len(data)} bytes)")
                # Try both with and without header skip
                #if len(data) == 20:
                    #raw = data
                if len(data) >= 23:
                    raw = data[3:23]
                else:
                    print(f"Unexpected data length: {len(data)}")
                    return None
                gas_names = ['CO', 'CO2', 'CH4', 'CnHm', 'H2', 'O2', 'C2H2', 'C2H4', 'HHV', 'N2']
                readings = {}
                for i, name in enumerate(gas_names):
                    value_bytes = raw[i*2:i*2+2]
                    value = int.from_bytes(value_bytes, byteorder='big', signed=False)
                    readings[name] = value / 100.0
                return readings
            else:
                print("Not enough bytes received from analyzer.")
                return None
        except Exception as e:
            print(f"Gas analyzer read error: {e}")
            return None







class DeviceData:
    def __init__(self):
        self.lock = threading.Lock()
        self.main_temps = [None] * 4
        self.ro_temps = [None] * 2
        self.setpoints = [25.0] * 4
        self.controller_states = [False] * 4
        self.controllers_enabled = [True, True, True, True]
        self.readonly_enabled = [True, True]
        self.pressures = ["--"] * 4
        self.power = None
        self.energy = None
        self.flow = None
        self.last_update = None
        self.mfc_flows = [None] * 4



def get_timestamped_filename(prefix="plot", ext="png"):
    now = datetime.now()

    timestamp = now.strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{ext}"

def get_data_type(col):
    name = col.lower()
    if any(x in name for x in ["temp", "heater"]):
        return "Temperature"
    if any(x in name for x in ["press", "psm4"]):
        return "Pressure"
    if any(x in name for x in ["flow", "mfc", "mfm"]):
        return "Flow"
    if "power" in name or "energy" in name:
        return "Power"
    if any(x in name for x in ["co", "o2", "h2", "ch4", "c2h2", "c2h4", "cn", "gas", "%"]):
        return "Gas"
    return "Other"

def find_most_recent_excel_file(folder_path):
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    if not files:
        return None
    full_paths = [os.path.join(folder_path, f) for f in files]
    most_recent_file = max(full_paths, key=os.path.getmtime)
    return most_recent_file

def plot_callback(user_data, x_axis_id, y_axis_ids):
    df = user_data['df']
    timestamps = df['Timestamp'].map(pd.Timestamp.timestamp)
    x = (timestamps - timestamps.iloc[0]).tolist()
    axes_with_data = {k: False for k in y_axis_ids}
    for axis_id in y_axis_ids.values():
        dpg.delete_item(axis_id, children_only=True)
    for col in user_data['series']:
        if col in df.columns:
            y = pd.to_numeric(df[col], errors='coerce').tolist()
            dtype = get_data_type(col)
            if dtype == "Temperature":
                axis_tag = "y_axis1"
            elif dtype == "Pressure":
                axis_tag = "y_axis2"
            else:
                axis_tag = "y_axis3"
            dpg.add_line_series(x, y, label=col, parent=y_axis_ids[axis_tag])
            axes_with_data[axis_tag] = True
    for axis_tag, axis_id in y_axis_ids.items():
        if axes_with_data[axis_tag]:
            dpg.show_item(axis_id)
            dpg.fit_axis_data(axis_id)
        else:
            dpg.hide_item(axis_id)

def update_plot(sender, app_data, user_data, x_axis_id, y_axis_ids):
    user_data['series'] = [col for col in user_data['available'] if dpg.get_value(f"chk_{col}")]
    plot_callback(user_data, x_axis_id, y_axis_ids)

def make_update_callback(user_data, x_axis_id, y_axis_ids):
    def callback(sender, app_data):
        update_plot(sender, app_data, user_data, x_axis_id, y_axis_ids)
    return callback

def save_plot_worker(user_data_dict, filename):
    import matplotlib
    matplotlib.use('Agg')  # Use non-GUI backend
    import matplotlib.pyplot as plt
    import pandas as pd

    df = pd.DataFrame(user_data_dict['df'])
    x = (pd.to_datetime(df['Timestamp']) - pd.to_datetime(df['Timestamp']).iloc[0]).dt.total_seconds()

    groups = {"Temperature": [], "Pressure": [], "Other": []}
    for col in user_data_dict['series']:
        dtype = None
        name = col.lower()
        if any(x in name for x in ["temp", "heater"]):
            dtype = "Temperature"
        elif any(x in name for x in ["press", "psm4"]):
            dtype = "Pressure"
        else:
            dtype = "Other"
        groups[dtype].append(col)

    fig, ax1 = plt.subplots(figsize=(12,6))
    color_cycle = plt.rcParams['axes.prop_cycle'].by_key()['color']

    if groups["Temperature"]:
        for i, col in enumerate(groups["Temperature"]):
            y = pd.to_numeric(df[col], errors='coerce')
            ax1.plot(x, y, label=col, color=color_cycle[i % len(color_cycle)])
        ax1.set_ylabel("Temperature (°C)")
        ax1.tick_params(axis='y')
    else:
        ax1.set_yticks([])

    ax2 = None
    if groups["Pressure"]:
        ax2 = ax1.twinx()
        for i, col in enumerate(groups["Pressure"]):
            y = pd.to_numeric(df[col], errors='coerce')
            ax2.plot(x, y, label=col,  color=color_cycle[(i+3) % len(color_cycle)])
        ax2.set_ylabel("Pressure (bar)")
        ax2.tick_params(axis='y')

    ax3 = None
    if groups["Other"]:
        ax3 = ax1.twinx()
        ax3.spines["right"].set_position(("axes", 1.1))
        for i, col in enumerate(groups["Other"]):
            y = pd.to_numeric(df[col], errors='coerce')
            ax3.plot(x, y, label=col,  color=color_cycle[(i+6) % len(color_cycle)])
        ax3.set_ylabel("Flow (slm) / Power (W, Wh) / Gas (%) / Other")
        ax3.tick_params(axis='y')

    ax1.set_xlabel("Time (s from start)")

    lines, labels = ax1.get_legend_handles_labels()
    if ax2:
        l2, lab2 = ax2.get_legend_handles_labels()
        lines += l2
        labels += lab2
    if ax3:
        l3, lab3 = ax3.get_legend_handles_labels()
        lines += l3
        labels += lab3
    if lines:
        ax1.legend(lines, labels, loc='best')

    plt.tight_layout()
    plt.savefig(filename, bbox_inches='tight')
    plt.close()


def save_plot_matplotlib(user_data):
    filename = get_timestamped_filename("plot")
    save_plot_worker(user_data, filename)
    print(f"Plot saved as {filename}")


def show_plot_window(log_file=None, blue_button_theme=None):
    if dpg.does_item_exist("history_plot_window"):
        dpg.delete_item("history_plot_window")

    if log_file is not None:
        file = log_file
    else:
        file = find_most_recent_excel_file(".")
    if file is None:
        print("No Excel files found in the folder.")
        return
    df = pd.read_excel(file)
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])
    available = [c for c in df.columns if c != "Timestamp"]
    user_data = {'df': df, 'series': [], 'available': available}
    default_selected = available[:1] if len(available) >= 3 else available
    user_data['series'] = default_selected
    with dpg.window(label="History Plot", width=1100, height=700, tag="history_plot_window"):

        with dpg.plot(label="History", height=400, width=1000, tag="history_plot"):
            dpg.add_plot_legend()
            x_axis_id = dpg.add_plot_axis(dpg.mvXAxis, label="Time (s from start)")
            y_axis1 = dpg.add_plot_axis(dpg.mvYAxis, label="Temperature (°C)", tag="y_axis1")
            y_axis2 = dpg.add_plot_axis(dpg.mvYAxis2, label="Pressure (bar)", tag="y_axis2")
            y_axis3 = dpg.add_plot_axis(dpg.mvYAxis3, label="Flow (slm) / Power (W, Wh) / Gas (%) / Other", tag="y_axis3")
            y_axis_ids = {"y_axis1": y_axis1, "y_axis2": y_axis2, "y_axis3": y_axis3}
        dpg.add_button(label="Save Plot", tag="save_plot_btn", callback=lambda s, a: save_plot_matplotlib(user_data))
        if blue_button_theme is not None:
            dpg.bind_item_theme("save_plot_btn", blue_button_theme)
        
           
        n = len(available)
        per_row = math.ceil(n / 2) if n > 8 else n
        for row in range(math.ceil(n / per_row)):
            with dpg.group(horizontal=True):
                for col in available[row*per_row:(row+1)*per_row]:
                    dpg.add_checkbox(
                        label=col,
                        tag=f"chk_{col}",
                        callback=make_update_callback(user_data, x_axis_id, y_axis_ids),
                        user_data=user_data,
                        default_value=(col in default_selected)
                    )
        plot_callback(user_data, x_axis_id, y_axis_ids)




class ControlApplication:
    def __init__(self):
        self.running = True
        # --- Data, settings, and devices ---
        self.button_command_queue = queue.Queue()
        self.serial_command_queue = queue.Queue()
        
        self.data = DeviceData()
        self.max_temp = 400.0
        self.max_press = 5.0
        self.load_settings()
        self.pause_polling_event = threading.Event()
        self.serial_command_queue = queue.Queue()
        
        self.mfc_states = [False]*4
        self.pre_emergency_heater_states = [False] * 4  # For 4 heaters
        self.pre_emergency_mfc_states = [False] * 4     # For 4 MFCs
        self.abnormal_logger = AbnormalEventLogger()

        # Devices
        self.pm = PowerMeter(port=PM_PORT)
        self.pm_enabled = True
        if self.pm_enabled:
            self.pm.connect()
        self.mfm = MFMFlowMeter(port=SERIAL_PORT, device_id=1)
        self.mfm_enabled = True
        self.logger = Logger(log_filename)
        self.mfc = MFCController(port=SERIAL_PORT)
        self.mfc_enabled = True
        self.gas_analyzer = GasAnalyzer(port=GAS_ANALYZER_PORT)
        self.gas_analyzer_enabled = True
        self.gas_values = {k: None for k in ['CO', 'CO2', 'CH4', 'CnHm', 'H2', 'O2', 'C2H2', 'C2H4', 'HHV', 'N2']}
        self.alarm_enabled = True
        self.modbus_client = ModbusSerialClient(
            port=SERIAL_PORT, baudrate=BAUDRATE, parity='N',
            stopbits=2, bytesize=8,timeout=0.2, retries=0)
        self.modbus_client.connect()
        self.tk4 = TK4Controller(self.modbus_client)
        self.psm4 = PSM4Controller(self.modbus_client)
        self.relay = ModbusRelayController(self.modbus_client, slave_id=8)

        # Start worker thread
        self.worker_thread = threading.Thread(target=self.serial_worker, daemon=True)
        self.worker_thread.start()

    # --- Settings ---
    def save_settings(self):
    # Load existing settings if present
        settings = {}
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r") as f:
                try:
                    settings = json.load(f)
                except Exception:
                    settings = {}

    # Update only the keys you want to change
        settings.update({
            "max_temp": self.max_temp,
            "max_press": self.max_press,
            "setpoints": self.data.setpoints,
            "mfc_setpoints": [dpg.get_value(f"mfc_set_{i}") for i in range(4)],
            "pre_emergency_heater_states": list(self.pre_emergency_heater_states),
            "pre_emergency_mfc_states": list(self.pre_emergency_mfc_states),
            "mfc_states": list(self.mfc_states),
        })

    # Write the updated dictionary back to the file, preserving all other keys
        with open(SETTINGS_FILE, "w") as f:
            json.dump(settings, f, indent=2)

    def load_settings(self):
        if not os.path.exists(SETTINGS_FILE):
            return
        with open(SETTINGS_FILE, "r") as f:
            settings = json.load(f)
        self.max_temp = settings.get("max_temp", self.max_temp)
        self.max_press = settings.get("max_press", self.max_press)
        self.data.setpoints = settings.get("setpoints", self.data.setpoints)
        self.mfc_setpoints = settings.get("mfc_setpoints", [0.0, 0.0, 0.0, 0.0])
        self.pre_emergency_heater_states = settings.get("pre_emergency_heater_states", [False]*4)
        self.pre_emergency_mfc_states = settings.get("pre_emergency_mfc_states", [False]*4)
        self.mfc_states = settings.get("mfc_states", [False]*4)


    # --- Serial worker thread ---
    def serial_worker(self):
        while self.running:
            try:
            # Always check button queue first!
                try:
                    command = self.button_command_queue.get_nowait()
                    self.process_command(command)
                    continue
                except queue.Empty:
                    pass

            # No urgent button command, check regular queue
                try:
                    command = self.serial_command_queue.get(timeout=0.1)
                    self.process_command(command)
                except queue.Empty:
                # No command: do periodic polling
                    self.handle_polling()
            except Exception as e:
                print(f"[Worker] Unexpected error: {e}")
            # Continue running even if an error occurs
                
    def process_command(self, command):
        cmd_type = command.get("cmd")
        reply_queue = command.get("reply_queue", None)

    # --- TK4 (Modbus) ---
        if cmd_type == "read_tk4":
            addr = command["address"]
            try:
                value = self.tk4.read_temperature(addr)
            except Exception as e:
                print(f"[Worker] read_tk4 error: {e}")
                value = None
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "address": addr, "value": value})

        elif cmd_type == "set_tk4_sv":
            addr = command["address"]
            val = command["value"]
            t0 = time.time()
            try:
                result = self.tk4.set_setpoint(addr, val)
            except Exception as e:
                print(f"[Worker] set_tk4_sv error: {e}")
                result = False
            t1 = time.time()
            print(f"[Worker] set_tk4_sv took {t1-t0:.2f}s")
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "address": addr, "success": result})

        elif cmd_type == "start_tk4":
            addr = command["address"]
            try:
                result = self.tk4.control_heater(addr, False)
            except Exception as e:
                print(f"start_tk4 error: {e}")
                result = False
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "address": addr, "success": result})

        

    # --- PSM4 (Modbus) ---
        elif cmd_type == "read_psm4":
            try:
                values = self.psm4.read_pressures()
            except Exception as e:
                print(f"[Worker] read_psm4 error: {e}")
                values = ["NC"] * 4
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "values": values})

    # --- Relay (Modbus) ---
        elif cmd_type == "relay_pulse":
            channel = command["channel"]
            duration = command.get("duration", 1.0)
            try:
                self.relay.send_pulse(channel, duration)
                success = True
            except Exception as e:
                print(f"[Worker] relay_pulse error: {e}")
                success = False
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "channel": channel, "success": success})

    # --- MFC (ASCII, needs port switch) ---
        elif cmd_type in ("set_mfc_flow", "on_off_mfc", "read_mfc"):
            try:
                self.modbus_client.close()
                if cmd_type == "set_mfc_flow":
                    ch = command["channel"]
                    val = command["value"]
                    try:
                        result = self.mfc.set_flow(ch, val)
                    except Exception as e:
                        print(f"[Worker] set_mfc_flow error: {e}")
                        result = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "channel": ch, "success": result})
                elif cmd_type == "on_off_mfc":
                    ch = command["channel"]
                    state = command["state"]
                    try:
                        result = self.mfc.on_off(ch, state)
                    except Exception as e:
                        print(f"[Worker] on_off_mfc error: {e}")
                        result = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "channel": ch, "success": result})
                elif cmd_type == "read_mfc":
                    ch = command["channel"]
                    try:
                        value = self.mfc.read_flow(ch)
                    except Exception as e:
                        print(f"[Worker] read_mfc error: {e}")
                        value = None
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "channel": ch, "value": value})
            finally:
                try:
                    self.modbus_client.connect()
                    self.tk4.client = self.modbus_client
                    self.psm4.client = self.modbus_client
                    self.relay.client = self.modbus_client
                except Exception as e:
                    print(f"[Worker] Error reconnecting Modbus after MFC: {e}")

    # --- MFM (ASCII, needs port switch) ---
        elif cmd_type == "read_mfm":
            try:
                self.modbus_client.close()
                try:
                    value = self.mfm.read_flow()
                except Exception as e:
                    print(f"[Worker] read_mfm error: {e}")
                    value = None
                if reply_queue:
                    reply_queue.put({"cmd": cmd_type, "value": value})
            finally:
                try:
                    self.modbus_client.connect()
                    self.tk4.client = self.modbus_client
                    self.psm4.client = self.modbus_client
                    self.relay.client = self.modbus_client
                except Exception as e:
                    print(f"[Worker] Error reconnecting Modbus after MFM: {e}")

    # --- Gas Analyzer (separate port) ---
        elif cmd_type == "read_gas":
            try:
                values = self.gas_analyzer.read_gases()
            except Exception as e:
                print(f"[Worker] read_gas error: {e}")
                values = None
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "values": values})

    # --- Poll all devices (periodic) ---
        elif cmd_type == "read_all":
            try:
                self.handle_polling()
            except Exception as e:
                print(f"[Worker] handle_polling error: {e}")
        
        elif cmd_type == "relay_open_all":
            try:
                self.relay.open_all()
                success = True
            except Exception as e:
                print(f"[Worker] relay_open_all error: {e}")
                success = False
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "success": success})

        elif cmd_type == "relay_close_all":
            try:
                self.relay.close_all()
                success = True
            except Exception as e:
                print(f"[Worker] relay_close_all error: {e}")
                success = False
            if reply_queue:
                reply_queue.put({"cmd": cmd_type, "success": success})



    def handle_polling(self):
    # 1. TK4 main controllers (poll one by one, yield to button queue between)
        for i, addr in enumerate(TK4_ADDRESSES):
            if self.data.controllers_enabled[i]:
                try:
                    self.data.main_temps[i] = self.tk4.read_temperature(addr)
                except Exception as e:
                    print(f"TK4 polling error for addr {addr}: {e}")
                    self.data.main_temps[i] = None
        # Yield to urgent button commands
            try:
                command = self.button_command_queue.get_nowait()
                self.process_command(command)
                return  # After processing, return to main worker loop
            except queue.Empty:
                pass

    # 2. TK4 read-only sensors
        for i, addr in enumerate(TK4_RO_ADDRESSES):
            if self.data.readonly_enabled[i]:
                try:
                    self.data.ro_temps[i] = self.tk4.read_temperature(addr)
                except Exception as e:
                    print(f"TK4 RO polling error for addr {addr}: {e}")
                    self.data.ro_temps[i] = None
            try:
                command = self.button_command_queue.get_nowait()
                self.process_command(command)
                return
            except queue.Empty:
                pass

    # 3. PSM4 pressures
        try:
            self.data.pressures = self.psm4.read_pressures()
        except Exception as e:
            print(f"PSM4 polling error: {e}")
            self.data.pressures = ["NC"] * 4
        try:
            command = self.button_command_queue.get_nowait()
            self.process_command(command)
            return
        except queue.Empty:
            pass
        
        if any(temp is not None and temp != 31000 and temp > self.max_temp for temp in self.data.main_temps):
            self.abnormal_logger.log_event("Max temperature exceeded")
            self.handle_emergency_stop()
            return

    # Check for over-pressure
        if any(
            press is not None and press != "NC" and press > self.max_press
            for press in self.data.pressures
        ):
            self.abnormal_logger.log_event("Max pressure exceeded")
            self.handle_emergency_stop()
            +eturn
        
    # 4. Power meter (if enabled)
        if self.pm_enabled:
            try:
                self.data.power = self.pm.read_power()
                self.data.energy = self.pm.read_energy()
            except Exception as e:
                print(f"Power meter polling error: {e}")
                self.data.power = None
                self.data.energy = None
            try:
                command = self.button_command_queue.get_nowait()
                self.process_command(command)
                return
            except queue.Empty:
                pass

    # 5. Close Modbus before MFM/MFC
        try:
            self.modbus_client.close()
        except Exception as e:
            print(f"Modbus close error: {e}")

    # 6. MFM
        if self.mfm_enabled:
            try:
                self.data.flow = self.mfm.read_flow()
            except Exception as e:
                print(f"MFM polling error: {e}")
                self.data.flow = None
            try:
                command = self.button_command_queue.get_nowait()
                self.process_command(command)
                return
            except queue.Empty:
                pass

    # 7. MFC
        if self.mfc_enabled:
            try:
                self.data.mfc_flows = self.mfc.read_all_flows()
            except Exception as e:
                print(f"MFC polling error: {e}")
                self.data.mfc_flows = [None] * 4
            try:
                command = self.button_command_queue.get_nowait()
                self.process_command(command)
                return
            except queue.Empty:
                pass

    # 8. Reopen Modbus after
        try:
            self.modbus_client.connect()
            self.tk4.client = self.modbus_client
            self.psm4.client = self.modbus_client
            self.relay.client = self.modbus_client
        except Exception as e:
            print(f"Modbus reconnect error: {e}")
        try:
            command = self.button_command_queue.get_nowait()
            self.process_command(command)
            return
        except queue.Empty:
            pass

    # 9. Gas analyzer
        if self.gas_analyzer_enabled and self.gas_analyzer:
            try:
                vals = self.gas_analyzer.read_gases()
                if vals and isinstance(vals, dict):
                    for k in self.gas_values:
                        self.gas_values[k] = vals.get(k)
            except Exception as e:
                print(f"Gas analyzer read error: {e}")
            try:
                command = self.button_command_queue.get_nowait()
                self.process_command(command)
                return
            except queue.Empty:
                pass

    # 10. Logging
        try:
            self.logger.log(self.data, self.gas_values)
        except Exception as e:
            print(f"Logger error: {e}")

        time.sleep(1)




        self.data.ro_temps = [self.tk4.read_temperature(addr) for addr in TK4_RO_ADDRESSES]
        self.data.pressures = self.psm4.read_pressures()
    # Close Modbus before MFC/MFM
        self.modbus_client.close()
        self.data.flow = self.mfm.read_flow()
        self.data.mfc_flows = self.mfc.read_all_flows()
    # Reopen Modbus after
        self.modbus_client.connect()
        self.tk4.client = self.modbus_client
        self.psm4.client = self.modbus_client
        self.relay.client = self.modbus_client
        if self.gas_analyzer_enabled and self.gas_analyzer:
            try:
                vals = self.gas_analyzer.read_gases()
                if vals and isinstance(vals, dict):
                    for k in self.gas_values:
                        self.gas_values[k] = vals.get(k)
            except Exception as e:
                print(f"Gas analyzer read error: {e}")
        self.logger.log(self.data, self.gas_values)

        time.sleep(1)


    # --- GUI callbacks use the queue ---
    def set_temperature(self, sender, app_data, user_data):
        index = user_data
        temperature = dpg.get_value(f"setpoint_{index}")
        reply_queue = queue.Queue()
        self.button_command_queue.put({
            "cmd": "set_tk4_sv",
            "address": TK4_ADDRESSES[index],
            "value": temperature,
            "reply_queue": reply_queue
        })
        try:
            result = reply_queue.get(timeout=2)
            if result.get("success"):
                self.data.setpoints[index] = temperature
                dpg.set_value(f"sv_display_{index}", f"{temperature:.1f}")
                dpg.set_value(f"setpoint_{index}", temperature)
                self.save_settings()
                self.update_status(f"Heater {index} set to {temperature:.1f}°C", COLOR_GREEN)
            else:
                self.update_status(f"Failed to set Heater {index+1}", COLOR_RED)
        except queue.Empty:
            self.update_status(f"Timeout: No response from Heater {index+1}", COLOR_RED)



    

    def stop_heater(self, sender, app_data, user_data):
        index = user_data
        stop_channel = index * 2 + 2
        self.button_command_queue.put({
            "cmd": "relay_pulse",
            "channel": stop_channel,
            "duration": 1.0
        })
        self.update_status(f"Heater {index+1} STOPPED (relay only)", COLOR_YELLOW)
        self.data.controller_states[index] = False



    def set_mfc_flow(self, sender, app_data, user_data):
        channel = user_data
        value = dpg.get_value(f"mfc_set_{channel}")
        reply_queue = queue.Queue()
        self.button_command_queue.put({
            "cmd": "set_mfc_flow",
            "channel": channel+1,
            "value": value,
            "reply_queue": reply_queue
        })
        result = reply_queue.get(timeout=2)
        if result.get("success"):
            self.mfc_setpoints[channel] = value
            dpg.set_value(f"mfc_sv_{channel}", f"{value:.2f}")
            dpg.set_value(f"mfc_set_{channel}", value)
            self.save_settings()
            self.update_status(f"Flow {channel+1} set")
        else:
            self.update_status(f"Flow {channel+1} set failed", COLOR_RED)

    def toggle_mfc_enabled(self, sender, app_data, user_data):
        import queue as pyqueue
        channel, state = user_data
        value = dpg.get_value(f"mfc_set_{channel}")
        status = "ON" if state else "OFF"

        try:
            if state:  # ON: set SV before turning on
                sv_queue = pyqueue.Queue()
                self.button_command_queue.put({
                    "cmd": "set_mfc_flow",
                    "channel": channel+1,
                    "value": value,
                    "reply_queue": sv_queue
                })
                sv_result = sv_queue.get(timeout=2)
                if sv_result.get("success"):
                    dpg.set_value(f"mfc_sv_{channel}", f"{value:.2f}")
                    dpg.set_value(f"mfc_set_{channel}", value)
                    self.mfc_setpoints[channel] = value
                    self.save_settings()
                else:
                    self.update_status(f"Flow {channel+1} SV set failed", COLOR_RED)
                    return

            reply_queue = pyqueue.Queue()
            self.button_command_queue.put({
                "cmd": "on_off_mfc",
                "channel": channel+1,
                "state": state,
                "reply_queue": reply_queue
            })
            result = reply_queue.get(timeout=2)
            if result.get("success"):
                self.mfc_states[channel] = state
                self.save_settings()
                self.update_status(f"Flow {channel+1} turned {status}", COLOR_GREEN if state else COLOR_YELLOW)
            else:
                self.update_status(f"Flow {channel+1} turn {status} failed", COLOR_RED)
        except pyqueue.Empty:
            self.update_status(f"Timeout: No response from Flow {channel+1} {status} command", COLOR_RED)

            
    def toggle_mfc_global(self, sender, app_data):
        self.mfc_enabled = app_data
        state = "enabled" if app_data else "disabled"
        self.update_status(f"Flow controller {state}", COLOR_GREEN if app_data else COLOR_YELLOW)
    
    def toggle_mfm_enabled(self, sender, app_data):
        self.mfm_enabled = app_data
        state = "enabled" if app_data else "disabled"
        self.update_status(f"MFM {state}", COLOR_GREEN if app_data else COLOR_YELLOW)



    def read_mfm_with_port_switch(self):
        """Safely read from the MFM (Mass Flow Meter) by temporarily switching serial port usage."""
        mfm_result = None
        try:
        # 1. Close the Modbus client to free the port
            if self.modbus_client:
                self.modbus_client.close()

        # 2. Open the serial port for MFM with correct parameters
            with serial.Serial(
                port=SERIAL_PORT,  # e.g., 'COM5'
                baudrate=9600,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                timeout=1
            ) as ser:
            # 3. Send MFM read command and read response
            # Replace the following with your actual MFM protocol commands
                ser.write(b'READ\r\n')
                mfm_result = ser.readline().decode().strip()

        except Exception as e:
            print(f"MFM read error: {e}")
            mfm_result = None

        finally:
        # 4. Reconnect the Modbus client for other devices
            try:
                if self.modbus_client:
                    self.modbus_client.connect()
                # Update client references in device classes if needed
                    self.tk4.client = self.modbus_client
                    self.psm4.client = self.modbus_client
                    self.relay.client = self.modbus_client
            except Exception as e:
                print(f"Error reconnecting Modbus after MFM: {e}")

        return mfm_result

    
    def toggle_gas_analyzer(self, sender, app_data):
        self.gas_analyzer_enabled = app_data
        state = "enabled" if app_data else "disabled"
        self.update_status(f"Gas analyzer {state}", COLOR_GREEN if app_data else COLOR_YELLOW)

    def open_settings_popup(self, sender, app_data):
        dpg.configure_item("settings_popup", show=True)

    def save_limits(self, sender, app_data):
        self.max_temp = dpg.get_value("max_temp_limit")
        self.max_press = dpg.get_value("max_press_limit")
        self.save_settings()
        dpg.configure_item("settings_popup", show=False)
        self.update_status(f"Limits set: Temp={self.max_temp}, Press={self.max_press}", COLOR_GREEN)

    def show_alarm(self, message):
        if not self.alarm_enabled:
            return
        dpg.set_value("alarm_text", message)
        dpg.configure_item("alarm_popup", show=True)
        dpg.configure_item("enable_alarm_btn", enabled=False)
        dpg.configure_item("disable_alarm_btn", enabled=True)
        try:
            import winsound
            winsound.PlaySound("mixkit-emergency-alert-alarm-1007.wav", winsound.SND_FILENAME | winsound.SND_ASYNC)
        except Exception:
            pass

    def enable_alarm(self, sender, app_data):
        self.alarm_enabled = True
        dpg.configure_item("enable_alarm_btn", enabled=False)
        dpg.configure_item("disable_alarm_btn", enabled=True)
        self.update_status("Alarms enabled", COLOR_GREEN)

    def disable_alarm(self, sender, app_data):
        self.alarm_enabled = False
        dpg.configure_item("enable_alarm_btn", enabled=True)
        dpg.configure_item("disable_alarm_btn", enabled=False)
    # Hide any current alarm popup and stop sound
        dpg.configure_item("alarm_popup", show=False)
        try:
            import winsound
            winsound.PlaySound(None, winsound.SND_PURGE)
        except Exception:
            pass
        self.update_status("Alarms disabled", COLOR_YELLOW)
     
    def handle_emergency_stop(self, sender=None, app_data=None):
    # Save current ON/OFF states before emergency stop
        self.abnormal_logger.log_event("Emergency stop activated")
        self.pre_emergency_heater_states = list(self.data.controller_states)
        self.pre_emergency_mfc_states = list(self.mfc_states)
        self.save_settings()
        

    # 1. Instantly open all relays
        relay_open_queue = queue.Queue()
        self.button_command_queue.put({
            "cmd": "relay_open_all",
            "reply_queue": relay_open_queue
        })
        try:
            relay_open_queue.get(timeout=2)
        except queue.Empty:
            print("[Emergency] Timeout opening all relays")

    # 2. Schedule closing all relays after 1 second (do NOT block GUI)
        def close_all_relays_later():
            time.sleep(1.0)
            relay_close_queue = queue.Queue()
            self.button_command_queue.put({
                "cmd": "relay_close_all",
                "reply_queue": relay_close_queue
            })
            try:
                relay_close_queue.get(timeout=2)
            except queue.Empty:
                print("[Emergency] Timeout closing all relays")
        threading.Thread(target=close_all_relays_later, daemon=True).start()

    # 3. Stop all MFCs (set flow to 0, turn off) - fire-and-forget, no wait
        for ch in self.mfc.channels:
            self.button_command_queue.put({
                "cmd": "set_mfc_flow",
                "channel": ch,
                "value": 0.0
            })
            self.button_command_queue.put({
                "cmd": "on_off_mfc",
                "channel": ch,
                "state": False
            })

    # 4. Show alarm/status in GUI
        self.show_alarm("EMERGENCY STOP ACTIVATED!\nAll relays toggled")
        self.update_status("EMERGENCY STOP: All relays toggled", COLOR_RED)
        self.abnormal_logger.log_event("Emergency stop activated")



 
    def handle_restart(self, sender=None, app_data=None):
        """Restore only those heaters and MFCs that were ON before Emergency Stop."""
    # 1. Restore heater setpoints and turn on only those previously ON
        for idx, addr in enumerate(TK4_ADDRESSES):
        # Restore setpoint
            sv = self.data.setpoints[idx]
            sv_queue = queue.Queue()
            self.button_command_queue.put({
                "cmd": "set_tk4_sv",
                "address": addr,
                "value": sv,
                "reply_queue": sv_queue
            })
            try:
                sv_queue.get(timeout=4)
            except queue.Empty:
                print(f"[Restart] Timeout restoring heater {idx+1} SV")
        # Only turn ON if it was ON before emergency
            if self.pre_emergency_heater_states[idx]:
                start_queue = queue.Queue()
                self.button_command_queue.put({
                    "cmd": "start_tk4",
                    "address": addr,
                    "reply_queue": start_queue
                })
                try:
                    result = start_queue.get(timeout=4)
                    if result.get("success"):
                        self.data.controller_states[idx] = True
                except queue.Empty:
                    print(f"[Restart] Timeout starting heater {idx+1}")

    # 2. Restore MFC flows and turn ON only those previously ON
        for i, ch in enumerate(self.mfc.channels):
            sv = dpg.get_value(f"mfc_set_{i}")
            set_queue = queue.Queue()
            self.button_command_queue.put({
                "cmd": "set_mfc_flow",
                "channel": ch,
                "value": sv,
                "reply_queue": set_queue
            })
            try:
                set_queue.get(timeout=2)
            except queue.Empty:
                print(f"[Restart] Timeout restoring MFC channel {ch} SV")
            if self.pre_emergency_mfc_states[i]:
                on_queue = queue.Queue()
                self.button_command_queue.put({
                    "cmd": "on_off_mfc",
                    "channel": ch,
                    "state": True,
                    "reply_queue": on_queue
                })
                try:
                    on_queue.get(timeout=2)
                except queue.Empty:
                    print(f"[Restart] Timeout turning ON MFC channel {ch}")

    # 3. Pulse relays for START as before
        for idx in range(4):
            if self.pre_emergency_heater_states[idx]:
                start_channel = idx * 2 + 1
                self.button_command_queue.put({
                    "cmd": "relay_pulse",
                    "channel": start_channel,
                    "duration": 1.0
                })

        self.update_status("RESTART: Operation resumed.", COLOR_GREEN)


 
           
    #GUI

    def create_gui(self):
        dpg.create_context()
        
        dpg.create_viewport(title='POX Process Control', width=1400, height=1200)

    # Fonts
        with dpg.font_registry():
            default_font = dpg.add_font("C:/Windows/Fonts/segoeui.ttf", 16)
            large_font = dpg.add_font("C:/Windows/Fonts/segoeui.ttf", 24)
            header_font = dpg.add_font("C:/Windows/Fonts/segoeuib.ttf", 20)
            alarm_font = dpg.add_font("C:/Windows/Fonts/segoeuib.ttf", 38)
        dpg.bind_font(default_font)

    # Button themes
        with dpg.theme() as blue_button_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Button, COLOR_BLUE)
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, [min(c+30, 255) for c in COLOR_BLUE])
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, [max(c-30, 0) for c in COLOR_BLUE])
        with dpg.theme() as green_button_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Button, COLOR_GREEN)
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, [min(c+30, 255) for c in COLOR_GREEN])
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, [max(c-30, 0) for c in COLOR_GREEN])
        with dpg.theme() as red_button_theme:
            with dpg.theme_component(dpg.mvButton):
                dpg.add_theme_color(dpg.mvThemeCol_Button, COLOR_RED)
                dpg.add_theme_color(dpg.mvThemeCol_ButtonHovered, [min(c+30, 255) for c in COLOR_RED])
                dpg.add_theme_color(dpg.mvThemeCol_ButtonActive, [max(c-30, 0) for c in COLOR_RED])

    # MAIN WINDOW
        with dpg.window(label="POX Control Dashboard", tag="primary_window"):
        # Top bar
            with dpg.group(horizontal=True):
                dpg.add_text("Last Update: Never", tag="last_update")
                button_width = 100  # Adjust as needed
                settings_btn = dpg.add_button(label="Settings", callback=self.open_settings_popup, width=button_width)
                enable_alarm_btn = dpg.add_button(label="Enable Alarm", tag="enable_alarm_btn", callback=self.enable_alarm, enabled=False, width=button_width)
                disable_alarm_btn = dpg.add_button(label="Disable Alarm", tag="disable_alarm_btn", callback=self.disable_alarm, enabled=True, width=button_width)
                plot_btn = dpg.add_button(
                label="Plot",
                callback=lambda: show_plot_window(blue_button_theme=blue_button_theme),
                width=button_width
            )


                dpg.add_spacer(width=10)
                dpg.add_text("", tag="status_text", color=(255, 255, 0))

                dpg.bind_item_theme(plot_btn, blue_button_theme)
                dpg.bind_item_theme(settings_btn, blue_button_theme)
                dpg.bind_item_theme(enable_alarm_btn, green_button_theme)
                dpg.bind_item_theme(disable_alarm_btn, red_button_theme)



            dpg.add_text("v1.0.4 | APGREEN", tag="corner_info", pos=(1200, 10), color=COLOR_GRAY)
            
            
            dpg.bind_item_font("corner_info", header_font)
            dpg.add_separator()

        # Main content: two columns
            with dpg.group(horizontal=True):
            # LEFT COLUMN
                with dpg.child_window(width=370, height=1200, tag="left_column"):
                    header_text = dpg.add_text("TEMPERATURE CONTROLLERS", color=COLOR_YELLOW)
                    dpg.bind_item_font(header_text, header_font)
                    dpg.add_separator()
                    heater_names = ["HEATER 1", "HEATER 2", "HEATER 3", "HEATER 4"]
                    for i in range(4):
                        with dpg.group():
                            h_text = dpg.add_text(heater_names[i], color=COLOR_WHITE)
                            dpg.bind_item_font(h_text, large_font)
                            with dpg.group(horizontal=True):
                                dpg.add_text("SV:", color=COLOR_GREEN)
                                sv_text = dpg.add_text("--.-", tag=f"sv_display_{i}", color=COLOR_GREEN)
                                dpg.bind_item_font(sv_text, large_font)
                                dpg.add_spacer(width=20)
                                dpg.add_text("PV:", color=COLOR_RED)
                                pv_text = dpg.add_text("--.-", tag=f"pv_display_{i}", color=COLOR_RED)
                                dpg.bind_item_font(pv_text, large_font)
                            dpg.add_checkbox(
                                label="Enabled", default_value=self.data.controllers_enabled[i],
                                callback=self.toggle_controller_enabled, tag=f"enable_{i}", user_data=i)
                            with dpg.group(horizontal=True):
                                dpg.add_input_float(label="Setpoint", tag=f"setpoint_{i}", default_value=self.data.setpoints[i], width=80, format="%.1f")
                                dpg.set_value(f"setpoint_{i}", self.data.setpoints[i])
                                set_btn = dpg.add_button(label="SET", callback=self.set_temperature, user_data=i, width=60, tag=f"set_btn_{i}")
                                dpg.bind_item_theme(set_btn, blue_button_theme)
                                start_btn = dpg.add_button(label="START", callback=self.start_heater, user_data=i, width=60, tag=f"start_btn_{i}")
                                dpg.bind_item_theme(start_btn, green_button_theme)
                                stop_btn = dpg.add_button(label="STOP", callback=self.stop_heater, user_data=i, width=60, tag=f"stop_btn_{i}")
                                dpg.bind_item_theme(stop_btn, red_button_theme)
                            dpg.add_separator()
                    checkpoint_names = ["TEMPERATURE 1", "TEMPERATURE 2"]
                    for i in range(2):
                        with dpg.group():
                            h_text = dpg.add_text(checkpoint_names[i], color=COLOR_WHITE)
                            dpg.bind_item_font(h_text, large_font)
                            with dpg.group(horizontal=True):
                                dpg.add_text("TEMP:", color=COLOR_WHITE)
                                ro_text = dpg.add_text("--.-", tag=f"readonly_temp_{i}", color=COLOR_WHITE)
                                dpg.bind_item_font(ro_text, large_font)
                            dpg.add_checkbox(label="Enabled", default_value=self.data.readonly_enabled[i],
                                         callback=self.toggle_readonly_enabled, tag=f"readonly_enable_{i}", user_data=i)
                            dpg.add_separator()

            # CENTER COLUMN
                with dpg.child_window(width=420, height=1200, tag="center_column"):
                    with dpg.group():
                        header = dpg.add_text("POWER METER", color=COLOR_YELLOW)
                        dpg.bind_item_font(header, header_font)
                        dpg.add_separator()
                        dpg.add_checkbox(label="Enabled", tag="pm_enabled", callback=self.toggle_pm_enabled, default_value=True)
                        with dpg.group(horizontal=True):
                            dpg.add_text("Power:", color=COLOR_GREEN)
                            power_text = dpg.add_text("0.00 W", tag="pm_power_display", color=COLOR_GREEN)
                            dpg.add_spacer(width=20)
                            dpg.add_text("Energy:", color=COLOR_BLUE)
                            energy_text = dpg.add_text("0.000 Wh", tag="pm_energy_display", color=COLOR_BLUE)
                            dpg.bind_item_font(power_text, large_font)
                            dpg.bind_item_font(energy_text, large_font)
                        with dpg.group(horizontal=True):
                            start_btn = dpg.add_button(label="Start", callback=self.pm_start, width=60)
                            dpg.bind_item_theme(start_btn, green_button_theme)
                            stop_btn = dpg.add_button(label="Stop", callback=self.pm_stop, width=60)
                            dpg.bind_item_theme(stop_btn, red_button_theme)
                            reset_btn = dpg.add_button(label="Reset", callback=self.pm_reset, width=60)
                            dpg.bind_item_theme(reset_btn, blue_button_theme)
                        dpg.add_separator()

                    header = dpg.add_text("PRESSURE INDICATOR", color=COLOR_YELLOW)
                    dpg.bind_item_font(header, header_font)
                    dpg.add_separator()
                    with dpg.group(horizontal=True):
                        for i in range(4):
                            dpg.add_checkbox(label=f"CH{i+1} Enabled", enabled=True, default_value=True, tag=f"psm_ch{i+1}_enabled")
                    with dpg.group(horizontal=True):
                        for i in range(4):
                            t = dpg.add_text("--", tag=f"psm_pressure_{i}", color=[0,150,255])
                            dpg.bind_item_font(t, large_font)
                            if i < 3:
                                dpg.add_spacer(width=60)
                    dpg.add_separator()

                    header = dpg.add_text("FLOWS", color=COLOR_YELLOW)
                    dpg.bind_item_font(header, header_font)
                    dpg.add_separator()
                    mfm_header = dpg.add_text("MFM", color=COLOR_YELLOW)
                    dpg.bind_item_font(mfm_header, header_font)
                    dpg.add_separator()
                    dpg.add_checkbox(label="MFM Enabled", default_value=self.mfm_enabled, callback=self.toggle_mfm_enabled, tag="mfm_enabled_checkbox")
                    with dpg.group(horizontal=True):
                        flow_text = dpg.add_text("--", tag="mfm_flow_display", color=COLOR_BLUE)
                        dpg.bind_item_font(flow_text, large_font)
                    dpg.add_separator()

                    header = dpg.add_text("MFC", color=COLOR_YELLOW)
                    dpg.bind_item_font(header, header_font)
                    dpg.add_separator()
                    with dpg.table(header_row=False):
                        for i in range(4):
                            dpg.add_table_column()
                        with dpg.table_row():
                            for i in range(4):
                                with dpg.table_cell():
                                    dpg.add_text(f"{i+1}", color=COLOR_WHITE)
                        with dpg.table_row():
                            for i in range(4):
                                with dpg.table_cell():
                                    dpg.add_text("SV", color=COLOR_GREEN)
                        with dpg.table_row():
                            for i in range(4):
                                with dpg.table_cell():
                                    sv_text = dpg.add_text("--", tag=f"mfc_sv_{i}", color=COLOR_GREEN)
                                    dpg.bind_item_font(sv_text, large_font)
                        with dpg.table_row():
                            for i in range(4):
                                with dpg.table_cell():
                                    dpg.add_text("PV", color=COLOR_RED)
                        with dpg.table_row():
                            for i in range(4):
                                with dpg.table_cell():
                                    pv_text = dpg.add_text("--", tag=f"mfc_pv_{i}", color=COLOR_RED)
                                    dpg.bind_item_font(pv_text, large_font)
                        with dpg.table_row():
                            for i in range(4):
                                with dpg.table_cell():
                                    dpg.add_input_float(width=80, format="%.3f", tag=f"mfc_set_{i}")
                                    if hasattr(self, "mfc_setpoints"):
                                        dpg.set_value(f"mfc_set_{i}", self.mfc_setpoints[i])
                        with dpg.table_row():
                            for i in range(4):
                                with dpg.table_cell():
                                    set_btn = dpg.add_button(label="SET", width=40, callback=self.set_mfc_flow, user_data=i)
                                    dpg.bind_item_theme(set_btn, blue_button_theme)
                                    on_btn = dpg.add_button(label="ON", width=40, callback=self.toggle_mfc_enabled, user_data=(i, True))
                                    dpg.bind_item_theme(on_btn, green_button_theme)
                                    off_btn = dpg.add_button(label="OFF", width=40, callback=self.toggle_mfc_enabled, user_data=(i, False))
                                    dpg.bind_item_theme(off_btn, red_button_theme)
                    dpg.add_checkbox(label="Flow Controller Enabled", default_value=self.mfc_enabled, callback=self.toggle_mfc_global)
                    
                    #RIGHT COLUMN
                    

                with dpg.child_window(width=600, height=1200, tag="right_column"):  # Increased width for 5 columns
                    header = dpg.add_text("GAS ANALYZER", color=COLOR_YELLOW)
                    dpg.bind_item_font(header, header_font)
                    dpg.add_separator()
                    with dpg.table(header_row=False):
                        # Add 5 columns for 5x2 layout
                        for _ in range(5):
                            dpg.add_table_column()
                        gases = ["CO", "CO2", "CH4", "CnHm", "H2", "O2", "C2H2", "C2H4", "HHV", "N2"]
        # First row: gases 0–4
                        with dpg.table_row():
                            for i in range(5):
                                gas = gases[i]
                                with dpg.table_cell():
                                    dpg.add_text(gas, color=COLOR_WHITE)
                                    gas_text = dpg.add_text("--.-", tag=f"gas_{gas}", color=COLOR_YELLOW)
                                    dpg.bind_item_font(gas_text, large_font)
        # Second row: gases 5–9
                        with dpg.table_row():
                            for i in range(5, 10):
                                gas = gases[i]
                                with dpg.table_cell():
                                    dpg.add_text(gas, color=COLOR_WHITE)
                                    gas_text = dpg.add_text("--.-", tag=f"gas_{gas}", color=COLOR_YELLOW)
                                    dpg.bind_item_font(gas_text, large_font)
                    dpg.add_checkbox(label="Gas Analyzer Enabled", tag="gas_analyzer_enabled", default_value=True, callback=self.toggle_gas_analyzer)
                    dpg.add_separator()
                    dpg.add_spacer(height=200)
                    dpg.add_button(
                        label="EMERGENCY STOP",
                        tag="emergency_btn",
                        width=600,
                        height=80,
                        callback=self.handle_emergency_stop,
                    )
                    dpg.bind_item_theme("emergency_btn", red_button_theme)
                    dpg.add_spacer(height=100)
                    dpg.add_button(
                        label="RESTART",
                        tag="restart_btn",
                        width=600,
                        height=60,
                        callback=self.handle_restart,
                    )
                    dpg.bind_item_theme("restart_btn", green_button_theme)

            

    # Settings popup window (modal)
        with dpg.window(label="Settings", modal=True, show=False, tag="settings_popup", no_title_bar=True, width=350, height=180):
            dpg.add_text("Set Maximum Limits")
            dpg.add_input_float(label="Max Temperature (°C)", tag="max_temp_limit", default_value=300.0,format="%.1f")
            dpg.set_value("max_temp_limit", self.max_temp)

            dpg.add_input_float(label="Max Pressure (bar)", tag="max_press_limit", default_value=5.0,format="%.1f")
            dpg.set_value("max_press_limit", self.max_press)

            with dpg.group(horizontal=True):
                dpg.add_button(label="Save", callback=self.save_limits)
                dpg.add_button(label="Cancel", callback=lambda: dpg.configure_item("settings_popup", show=False))
        with dpg.window(label="Alarm", modal=True, show=False, tag="alarm_popup", no_title_bar=True, width=500, height=180):
            alarm_text = dpg.add_text("", tag="alarm_text", color=[255, 0, 0])
            dpg.bind_item_font(alarm_text, alarm_font)
            dpg.add_spacer(height=10)
            dpg.add_button(label="OK", width=120, height=40, callback=lambda: dpg.configure_item("alarm_popup", show=False))
        for i in range(4):
            dpg.set_value(f"sv_display_{i}", f"{self.data.setpoints[i]:.1f}")
            dpg.set_value(f"setpoint_{i}", self.data.setpoints[i])
        if hasattr(self, "mfc_setpoints"):
            for i in range(4):
                dpg.set_value(f"mfc_set_{i}", self.mfc_setpoints[i])
                dpg.set_value(f"mfc_sv_{i}", f"{self.mfc_setpoints[i]:.2f}")
        
        dpg.setup_dearpygui()
        dpg.show_viewport()
        dpg.set_primary_window("primary_window", True)

    
    def update_status(self, message, color=COLOR_WHITE):
        """Update status message with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        dpg.set_value("status_text", f"[{timestamp}] {message}")
        dpg.configure_item("status_text", color=color)
    
    def toggle_controller_enabled(self, sender, app_data, user_data):
        """Toggle controller enabled state"""
        index = user_data
        self.data.controllers_enabled[index] = app_data
        state = "enabled" if app_data else "disabled"
        self.update_status(f"Heater {index+1} {state}", 
                         COLOR_GREEN if app_data else COLOR_YELLOW)
    
    def toggle_readonly_enabled(self, sender, app_data, user_data):
        """Toggle read-only sensor enabled state"""
        index = user_data
        self.data.readonly_enabled[index] = app_data
        state = "enabled" if app_data else "disabled"
        self.update_status(f"Checkpoint {index+1} {state}", 
                         COLOR_GREEN if app_data else COLOR_YELLOW)
    
    def toggle_pm_enabled(self, sender, app_data):
        """Toggle power meter enabled state"""
        self.pm_enabled = app_data
        try:
            if app_data:
                success = self.pm.connect()
                if success:
                    self.update_status("Power meter connected", COLOR_GREEN)
                else:
                    self.update_status("Power meter connection failed", COLOR_RED)
            else:
                self.pm.close()
                self.update_status("Power meter disconnected", COLOR_YELLOW)
        except Exception as e:
            self.update_status(f"Power meter error: {str(e)}", COLOR_RED)
    
    def set_temperature(self, sender, app_data, user_data):
        index = user_data
        temperature = dpg.get_value(f"setpoint_{index}")
        reply_queue = queue.Queue()
        self.button_command_queue.put({
            "cmd": "set_tk4_sv",
            "address": TK4_ADDRESSES[index],
            "value": temperature,
            "reply_queue": reply_queue
        })
        result = reply_queue.get(timeout=2)
        if result.get("success"):
            self.data.setpoints[index] = temperature
            dpg.set_value(f"sv_display_{index}", f"{temperature:.1f}")
            dpg.set_value(f"setpoint_{index}", temperature)
            self.save_settings()
            self.update_status(f"Heater {index+1} set to {temperature:.1f}°C", COLOR_GREEN)
        else:
            self.update_status(f"Failed to set Heater {index+1}", COLOR_RED)


    def start_heater(self, sender, app_data, user_data):
        index = user_data
        temperature = dpg.get_value(f"setpoint_{index}")
        sv_queue = queue.Queue()
        self.button_command_queue.put({
            "cmd": "set_tk4_sv",
            "address": TK4_ADDRESSES[index],
            "value": temperature,
            "reply_queue": sv_queue
        })
        try:
            sv_result = sv_queue.get(timeout=4)
            if sv_result.get("success"):
                dpg.set_value(f"sv_display_{index}", f"{temperature:.1f}")
                self.data.setpoints[index] = temperature
                self.save_settings()
            else:
                self.update_status(f"Failed to set Heater {index+1}", COLOR_RED)
                return
        except queue.Empty:
            self.update_status(f"Timeout: No response from Heater {index+1} setpoint", COLOR_RED)
            return

        reply_queue = queue.Queue()
        self.button_command_queue.put({
            "cmd": "start_tk4",
            "address": TK4_ADDRESSES[index],
            "reply_queue": reply_queue
        })
        try:
            result = reply_queue.get(timeout=4)
            if result.get("success"):
                self.update_status(f"Heater {index+1} RUNNING", COLOR_GREEN)
                self.data.controller_states[index] = True  # <-- THIS IS NEEDED!
    # Only pulse relay if start succeeded
                start_channel = index * 2 + 1
                print(f"Queuing relay_pulse for START: channel={start_channel}")
                self.button_command_queue.put({
                    "cmd": "relay_pulse",
                    "channel": start_channel,
                    "duration": 1.0
                })  
            else:
                self.update_status(f"Failed to start Heater {index+1}", COLOR_RED)

        except queue.Empty:
            self.update_status(f"Timeout: No response from Heater {index+1} start", COLOR_RED)


    def stop_heater(self, sender, app_data, user_data):
        index = user_data
        stop_channel = index * 2 + 2
        self.button_command_queue.put({
            "cmd": "relay_pulse",
            "channel": stop_channel,
            "duration": 1.0
        })
        self.update_status(f"Heater {index+1} STOPPED (relay only)", COLOR_YELLOW)
        self.data.controller_states[index] = False






    
    def pm_start(self):
        """Start power meter integration"""
        if not self.pm_enabled:
            self.update_status("Power meter not enabled", COLOR_YELLOW)
            return
        
        try:
            self.pm.start_integration()
            self.update_status("Power meter integration started", COLOR_GREEN)
        except Exception as e:
            self.update_status(f"Power meter error: {str(e)}", COLOR_RED)
    
    def pm_stop(self):
        """Stop power meter integration"""
        if not self.pm_enabled:
            self.update_status("Power meter not enabled", COLOR_YELLOW)
            return
        
        try:
            self.pm.stop_integration()
            self.update_status("Power meter integration stopped", COLOR_YELLOW)
        except Exception as e:
            self.update_status(f"Power meter error: {str(e)}", COLOR_RED)
    
    def pm_reset(self):
        """Reset power meter integration"""
        if not self.pm_enabled:
            self.update_status("Power meter not enabled", COLOR_YELLOW)
            return
        
        try:
            self.pm.reset_integration()
            self.update_status("Power meter integration reset", COLOR_BLUE)
        except Exception as e:
            self.update_status(f"Power meter error: {str(e)}", COLOR_RED)
    
    def refresh_all(self):
        """Manually refresh all displays"""
        self.update_status("Manual refresh initiated", COLOR_BLUE)
        self.update_displays()
        self.update_status("Refresh completed", COLOR_GREEN)
    
    def update_displays(self):
    # TK4 Controllers
        for i in range(4):
            dpg.set_value(f"sv_display_{i}", f"{self.data.setpoints[i]:.1f}")  # <-- Add this line
            if self.data.controllers_enabled[i]:
                temp = self.data.main_temps[i]
                if temp is not None:
                    if temp > 2000:
                        dpg.set_value(f"pv_display_{i}", "OPEN")
                    else:
                        dpg.set_value(f"pv_display_{i}", f"{temp:.1f}")
                else:
                    dpg.set_value(f"pv_display_{i}", "--.-")
            else:
                dpg.set_value(f"pv_display_{i}", "--.-")
    #

        # Read-only sensors
        for i in range(2):
            if self.data.readonly_enabled[i]:
                temp = self.data.ro_temps[i]
                if temp is not None:
                    if temp > 2000:
                        dpg.set_value(f"readonly_temp_{i}", "OPEN")
                    else:
                        dpg.set_value(f"readonly_temp_{i}", f"{temp:.1f}")
                else:
                    dpg.set_value(f"readonly_temp_{i}", "--.-")
            else:
                dpg.set_value(f"readonly_temp_{i}", "--.-")
                
                    
        # Update PSM4 pressures
        for i in range(4):
            dpg.set_value(f"psm_pressure_{i}", self.data.pressures[i])
        # Power meter
        if self.pm_enabled:
            power = self.data.power
            energy = self.data.energy
            dpg.set_value("pm_power_display", f"{power:.2f} W" if power is not None else "0.00 W")
            dpg.set_value("pm_energy_display", f"{energy:.3f} Wh" if energy is not None else "0.000 Wh")
        else:
            dpg.set_value("pm_power_display", "0.00 W")
            dpg.set_value("pm_energy_display", "0.000 Wh")
        flow = self.data.flow
        if flow is not None:
            dpg.set_value("mfm_flow_display", f"{flow:.2f}")
        else:
            dpg.set_value("mfm_flow_display", "--")

        for gas in ['CO', 'CO2', 'CH4', 'CnHm', 'H2', 'O2', 'C2H2', 'C2H4', 'HHV', 'N2']:
            val = self.gas_values.get(gas)
            dpg.set_value(f"gas_{gas}", f"{val:.2f}" if val is not None else "--.-")


        # Last update time
        dpg.set_value("last_update", f"Last Update: {datetime.now().strftime('%H:%M:%S')}")
        for i in range(4):
            val = self.data.mfc_flows[i]
            dpg.set_value(f"mfc_pv_{i}", f"{val:.2f}" if val is not None else "--")
    
    def run(self):
        last_update = 0
        update_interval = 0.05
        while dpg.is_dearpygui_running():
            current_time = time.time()
            if current_time - last_update >= update_interval:
                self.update_displays()
                last_update = current_time
            dpg.render_dearpygui_frame()
        self.running = False
        self.worker_thread.join(timeout=2)
        if self.pm_enabled:
            self.pm.close()
        self.tk4.close()
        self.save_settings()
        dpg.destroy_context()



if __name__ == "__main__":
    app = ControlApplication()
    app.create_gui()
    app.run()

        

