import wx
import threading
import queue
import json
import os
import datetime
import matplotlib
matplotlib.use('WXAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.figure import Figure
from pymodbus.client import ModbusSerialClient
#import sys
import winsound
import serial
import gc
import struct
import time
wx.Log.SetActiveTarget(wx.LogStderr())
import logging
import openpyxl
from openpyxl.utils import get_column_letter
import glob
import pandas as pd
import serial.tools.list_ports


# Get parent directory of the script
script_dir = os.path.dirname(os.path.abspath(__file__))
data_log_dir = os.path.join(script_dir, "Data log")

# Create the folder if it doesn't exist
os.makedirs(data_log_dir, exist_ok=True)

# Set log filename in that folder
log_filename = os.path.join(
    data_log_dir,
    f"process_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

def get_data_type(col):
    name = col.lower()
    if any(x in name for x in ["temp", "heater"]):
        return "Temperature"
    if any(x in name for x in ["press", "psm4"]):
        return "Pressure"
    if any(x in name for x in ["flow", "mfc", "mfm"]):
        return "Flow"
    if "power" in name or "energy" in name:
        return "Power"
    if any(x in name for x in ["co", "o2", "h2", "ch4", "c2h2", "c2h4", "cn", "gas", "%"]):
        return "Gas"
    return "Other"



class PlotDialog(wx.Dialog):
    def __init__(self, parent, log_file, data_log_dir):
        super().__init__(parent, title="Data Plot", size=(1200, 800))
        self.SetBackgroundColour(wx.Colour(0, 0, 0))  # All-black window

        self.log_file = log_file
        self.data_log_dir = data_log_dir

        panel = wx.Panel(self)
        panel.SetBackgroundColour(wx.Colour(0, 0, 0))
        vbox = wx.BoxSizer(wx.VERTICAL)

        # Load data
        self.df = pd.read_excel(self.log_file)
        self.df['Timestamp'] = pd.to_datetime(self.df['Timestamp'])
        self.available = [c for c in self.df.columns if c != "Timestamp"]

        # --- Checkbox panel for series selection ---
        self.checkboxes = {}
        checkbox_panel = wx.Panel(panel)
        checkbox_panel.SetBackgroundColour(wx.Colour(0, 0, 0))
        grid = wx.GridSizer(rows=0, cols=7, hgap=5, vgap=5)
        for col in self.available:
            cb = wx.CheckBox(checkbox_panel, label=col, size=(20, 20))
            cb.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_NORMAL))
            cb.SetForegroundColour(wx.Colour(255, 255, 255))
            cb.SetBackgroundColour(wx.Colour(0, 0, 0))
            cb.SetValue(False)
            cb.Bind(wx.EVT_CHECKBOX, self.on_checkbox)
            self.checkboxes[col] = cb
            grid.Add(cb, flag=wx.EXPAND)
        checkbox_panel.SetSizer(grid)
        vbox.Add(checkbox_panel, flag=wx.EXPAND | wx.ALL, border=10)

        # --- Matplotlib Figure ---
        matplotlib.rcParams['axes.edgecolor'] = 'white'
        matplotlib.rcParams['axes.labelcolor'] = 'white'
        matplotlib.rcParams['xtick.color'] = 'white'
        matplotlib.rcParams['ytick.color'] = 'white'
        matplotlib.rcParams['text.color'] = 'white'
        matplotlib.rcParams['figure.facecolor'] = 'black'
        matplotlib.rcParams['axes.facecolor'] = 'black'
        matplotlib.rcParams['savefig.facecolor'] = 'black'
        self.figure = Figure(figsize=(12, 5), facecolor='black')
        self.canvas = FigureCanvas(panel, -1, self.figure)
        vbox.Add(self.canvas, 1, flag=wx.EXPAND | wx.ALL, border=10)

        # --- Save/Close buttons ---
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        font = wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        self.save_btn = wx.Button(panel, label="Save Plot", size=(160, 50))
        self.save_btn.SetFont(font)
        self.save_btn.Bind(wx.EVT_BUTTON, self.on_save_plot)
        hbox.Add(self.save_btn, flag=wx.RIGHT, border=10)
        close_btn = wx.Button(panel, label="Close", size=(160, 50))
        close_btn.SetFont(font)
        close_btn.Bind(wx.EVT_BUTTON, lambda evt: self.EndModal(wx.ID_OK))
        hbox.Add(close_btn)
        vbox.Add(hbox, flag=wx.ALIGN_RIGHT | wx.ALL, border=10)

        panel.SetSizer(vbox)
        self.selected = set()
        self.plot_data()

    def on_checkbox(self, event):
        self.selected = set(col for col, cb in self.checkboxes.items() if cb.GetValue())
        self.plot_data()

    def plot_data(self):
        self.figure.clear()
        t0 = self.df['Timestamp'].iloc[0]
        x = (self.df['Timestamp'] - t0).dt.total_seconds()

        # Group selected columns by type
        groups = {"Temperature": [], "Pressure": [], "Other": []}
        for col in self.selected:
            dtype = get_data_type(col)
            if dtype == "Temperature":
                groups["Temperature"].append(col)
            elif dtype == "Pressure":
                groups["Pressure"].append(col)
            else:
                groups["Other"].append(col)

        # Create up to 3 Y axes
        ax1 = self.figure.add_subplot(111)
        ax2 = ax1.twinx() if groups["Pressure"] or groups["Other"] else None
        ax3 = None
        if groups["Pressure"] and groups["Other"]:
            ax3 = ax1.twinx()
            ax3.spines['right'].set_position(('axes', 1.15))
            ax3.spines['right'].set_visible(True)

        color_cycle = plt.rcParams['axes.prop_cycle'].by_key()['color']
        lines = []
        labels = []

        # Temperature
        for i, col in enumerate(groups["Temperature"]):
            y = pd.to_numeric(self.df[col], errors='coerce')
            l, = ax1.plot(x, y, label=col, linewidth=2)
            lines.append(l)
            labels.append(col)
        ax1.set_ylabel("Temperature (°C)", color="white")
        ax1.tick_params(axis='y', labelcolor="white")

        # Pressure
        if groups["Pressure"]:
            for i, col in enumerate(groups["Pressure"]):
                y = pd.to_numeric(self.df[col], errors='coerce')
                l, = ax2.plot(x, y, label=col, linewidth=2)
                lines.append(l)
                labels.append(col)
            ax2.set_ylabel("Pressure (bar)", color="white")
            ax2.tick_params(axis='y', labelcolor="white")

        # Other (Flow, Power, Gas, etc.)
        if groups["Other"]:
            target_ax = ax3 if ax3 else (ax2 if ax2 else ax1)
            for i, col in enumerate(groups["Other"]):
                y = pd.to_numeric(self.df[col], errors='coerce')
                l, = target_ax.plot(x, y, label=col, linewidth=2)
                lines.append(l)
                labels.append(col)
            if ax3:
                ax3.set_ylabel("Other", color="white")
                ax3.tick_params(axis='y', labelcolor="white")
            elif ax2 and not groups["Pressure"]:
                ax2.set_ylabel("Other", color="white")
                ax2.tick_params(axis='y', labelcolor="white")

        ax1.set_xlabel("Time (s from start)", color="white")
        ax1.tick_params(axis='x', labelcolor="white")
        if lines:
            legend = ax1.legend(lines, labels, loc='upper left', fontsize=12)
            if legend:
                for text in legend.get_texts():
                    text.set_color("white")
        self.figure.tight_layout()
        self.canvas.draw()

    def save_full_plot(self):
        # Use a dark background for the saved plot
        plt.style.use('dark_background')
        matplotlib.rcParams['axes.edgecolor'] = 'white'
        matplotlib.rcParams['axes.labelcolor'] = 'white'
        matplotlib.rcParams['xtick.color'] = 'white'
        matplotlib.rcParams['ytick.color'] = 'white'
        matplotlib.rcParams['text.color'] = 'white'
        matplotlib.rcParams['figure.facecolor'] = 'black'
        matplotlib.rcParams['axes.facecolor'] = 'black'
        matplotlib.rcParams['savefig.facecolor'] = 'black'

        df = self.df
        t0 = df['Timestamp'].iloc[0]
        x = (df['Timestamp'] - t0).dt.total_seconds()

    # Group selected columns by type
        groups = {"Temperature": [], "Pressure": [], "Other": []}
        for col in self.selected:
            dtype = get_data_type(col)
            if dtype == "Temperature":
                groups["Temperature"].append(col)
            elif dtype == "Pressure":
                groups["Pressure"].append(col)
            else:
                groups["Other"].append(col)

        fig, ax1 = plt.subplots(figsize=(12, 5))
        ax2 = ax1.twinx() if groups["Pressure"] or groups["Other"] else None
        ax3 = None
        if groups["Pressure"] and groups["Other"]:
            ax3 = ax1.twinx()
            ax3.spines['right'].set_position(('axes', 1.15))
            ax3.spines['right'].set_visible(True)

        color_cycle = plt.rcParams['axes.prop_cycle'].by_key()['color']
        lines = []
        labels = []

    # Temperature
        for i, col in enumerate(groups["Temperature"]):
            y = pd.to_numeric(df[col], errors='coerce')
            l, = ax1.plot(x, y, label=col, linewidth=2)
            lines.append(l)
            labels.append(col)
        ax1.set_ylabel("Temperature (°C)", color="white")
        ax1.tick_params(axis='y', labelcolor="white")

    # Pressure
        if groups["Pressure"]:
            for i, col in enumerate(groups["Pressure"]):
                y = pd.to_numeric(df[col], errors='coerce')
                l, = ax2.plot(x, y, label=col, linewidth=2)
                lines.append(l)
                labels.append(col)
            ax2.set_ylabel("Pressure (bar)", color="white")
            ax2.tick_params(axis='y', labelcolor="white")

    # Other (Flow, Power, Gas, etc.)
        if groups["Other"]:
            target_ax = ax3 if ax3 else (ax2 if ax2 else ax1)
            for i, col in enumerate(groups["Other"]):
                y = pd.to_numeric(df[col], errors='coerce')
                l, = target_ax.plot(x, y, label=col, linewidth=2)
                lines.append(l)
                labels.append(col)
            if ax3:
                ax3.set_ylabel("Other", color="white")
                ax3.tick_params(axis='y', labelcolor="white")
            elif ax2 and not groups["Pressure"]:
                ax2.set_ylabel("Other", color="white")
                ax2.tick_params(axis='y', labelcolor="white")

        ax1.set_xlabel("Time (s from start)", color="white")
        ax1.tick_params(axis='x', labelcolor="white")
        if lines:
            legend = ax1.legend(lines, labels, loc='upper left', fontsize=12)
            if legend:
                for text in legend.get_texts():
                    text.set_color("white")
        fig.tight_layout()

        now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.join(self.data_log_dir, f"plot_{now}.png")
        plt.savefig(filename, bbox_inches='tight')
        plt.close(fig)
        return filename


    def on_save_plot(self, event):
        filename = self.save_full_plot()
        wx.MessageBox(f"Plot saved to {filename}", "Save Plot", wx.ICON_INFORMATION)


class Logger:
    def __init__(self):
        self.columns = [
            "Timestamp", "Heater", "Preheater", "Reactor",
            "Temp1", "Temp2", "Temp3",
            "Pressure1", "Pressure2", "Pressure3",
            "Power", "Energy", "MFM_Flow",
            "MFC_CH4", "MFC_O2", "MFC_N2", "MFC_H2",
            "CO", "CO2", "CH4", "CnHm", "H2", "O2", "C2H2", "C2H4", "HHV", "N2"
        ]
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.data_log_dir = os.path.join(script_dir, "Data log")
        os.makedirs(self.data_log_dir, exist_ok=True)

    def get_filename(self):
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H')
        return os.path.join(self.data_log_dir, f"process_log_{timestamp}.xlsx")

    def log(self, data, gas_values=None):
        filename = self.get_filename()

        # Create file and header if new
        if not os.path.exists(filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(self.columns)
            ws.column_dimensions['A'].width = 18
            for i, col in enumerate(self.columns[1:], 2):
                ws.column_dimensions[get_column_letter(i)].width = 12
            wb.save(filename)

        row = [
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            data.get(1), data.get(2), data.get(3),
            data.get(4), data.get(5), data.get(6),
            *(data.get('pressures', [None, None, None])),
            data.get('power'), data.get('energy'), data.get('mfm_flow'),
            *(data.get('mfc_flows', [None, None, None, None])),
            *(gas_values.get(gas) if gas_values else None for gas in ["CO", "CO2", "CH4", "CnHm", "H2", "O2", "C2H2", "C2H4", "HHV", "N2"])
        ]

        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            ws.append(row)
            wb.save(filename)
        except Exception as e:
            print(f"[Logger] Logging error: {e}")







abnormal_logger = logging.getLogger("abnormal")
abnormal_logger.setLevel(logging.INFO)
abnormal_handler = logging.FileHandler("abnormal_events.log")
abnormal_handler.setFormatter(logging.Formatter('%(asctime)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
abnormal_logger.handlers = [abnormal_handler]  # Remove all other handlers

def log_abnormal_event(description):
    abnormal_logger.info(description)

# Use an absolute path for the settings file
SETTINGS_FILE = os.path.join(os.path.dirname(__file__), "settings.json")

# Default settings for the JSON file
DEFAULT_SETTINGS = {
    "sv": "--",
    "max_temp": "--",
    "sensor_settings": {
        "sensor_1_max_pressure": 20
    },
    "pressure_values": {
        "sensor_1": "-- bar"
    },
    "heater_2": {
        "coil_sv": "--",
        "coil_pv": "--",
        "reactor_temp": "--",
        "coil_max_temp": "--",
        "reactor_max_temp": "--",
        "status": "OFF"
    }
}

# Ensure the JSON file exists and create it with default values if missing
if not os.path.exists(SETTINGS_FILE):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(DEFAULT_SETTINGS, f, indent=4)


    
def update_datetime(self):
    """Update the date and time in the status bar."""
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    self.datetime_label.SetLabel(now)
    wx.CallLater(1000, self.update_datetime)  # Update every second

def start_datetime_updates(self):
    """Start periodic updates for the date and time."""
    self.update_datetime()

class NumericKeypad(wx.Dialog):
    def __init__(self, parent, title="Enter Value"):
        super().__init__(parent, title=title, size=(300, 400))

        self.value = ""  # Store the entered value

        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)

        # Display area for the entered value
        self.display = wx.TextCtrl(panel, style=wx.TE_READONLY | wx.TE_CENTER, size=(250, 50))
        vbox.Add(self.display, flag=wx.ALIGN_CENTER | wx.TOP, border=10)

        # Create numeric keypad buttons
        grid = wx.GridSizer(5, 3, 5, 5)  # 4 rows, 3 columns
        for label in ["1", "2", "3", "4", "5", "6", "7", "8", "9", ".", "0", "Backspace", "Clear", "OK"]:

            button = wx.Button(panel, label=label, size=(70, 50))
            grid.Add(button, flag=wx.EXPAND)
            button.Bind(wx.EVT_BUTTON, self.on_button_click)

        vbox.Add(grid, flag=wx.ALIGN_CENTER | wx.ALL, border=10)
        panel.SetSizer(vbox)

    def on_button_click(self, event):
        label = event.GetEventObject().GetLabel()
        if label == "Backspace":
            self.value = self.value[:-1]
        if label == "Clear":
            self.value = ""
        elif label == "OK":
            self.EndModal(wx.ID_OK)
            return
        elif label == ".":
            if "." not in self.value:
                self.value += "."
        elif label.isdigit():
            self.value += label
        self.display.SetValue(self.value)

    def get_value(self):
        """Return the entered value."""
        return self.value



class ModbusRelayController:
    """
    Controls up to 4 relays via Modbus RTU relay module (e.g., Waveshare, KMTronic).
    Uses write_register (not write_coil!) for relay commands.
    """
    def __init__(self, client, lock=None, slave_id=8):
        self.client = client
        self.slave_id = slave_id
        self.lock = lock or threading.Lock()

    def _write_channel(self, channel, command):
        value = 0x0100 if command == 'on' else 0x0200
        with self.lock:
        # Always reconnect before every relay command!
            if not self.client.is_socket_open():
                try:
                    self.client.connect()
                    time.sleep(0.1)  # Give the OS/driver a moment
                except Exception as e:
                    print(f"Relay: failed to connect before command: {e}")
            try:
                self.client.write_register(channel, value, slave=self.slave_id)
            except Exception as e:
                print(f"Relay: write_register error: {e}")


    def send_pulse(self, channel, duration=1.0):
        try:
            self._write_channel(channel - 1, 'on')
            time.sleep(duration)
            self._write_channel(channel - 1, 'off')
        except Exception as e:
            print(f"Relay pulse error on channel {channel}: {e}")


    def open_all(self):
        """Open (activate) all relays using the group command."""
        with self.lock:
            if not self.client.is_socket_open():
                self.client.connect()
            self.client.write_register(0x0000, 0x0700, slave=self.slave_id)

    def close_all(self):
        """Close (deactivate) all relays using the group command."""
        with self.lock:
            if not self.client.is_socket_open():
                self.client.connect()
            self.client.write_register(0x0000, 0x0800, slave=self.slave_id)



class HeaterDialog(wx.Dialog):
    def __init__(self, parent, apply_changes_callback, settings):
        super().__init__(parent, title="Preheater Settings", size=(300, 250))
        self.apply_changes = apply_changes_callback
        self.settings = settings

        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)

        vbox.Add(wx.StaticText(panel, label="Set Temperature (SV):"), flag=wx.LEFT | wx.TOP, border=10)
        self.temp_input = wx.TextCtrl(panel, value=str(self.settings.get("heater_1", {}).get("sv", "--")))
        self.temp_input.Bind(wx.EVT_LEFT_DOWN, self.show_keypad)
        vbox.Add(self.temp_input, flag=wx.LEFT | wx.EXPAND, border=10)

        vbox.Add(wx.StaticText(panel, label="Set Maximum Temperature:"), flag=wx.LEFT | wx.TOP, border=10)
        self.max_temp_input = wx.TextCtrl(panel, value=str(self.settings.get("heater_1", {}).get("max_temp", "--")))
        self.max_temp_input.Bind(wx.EVT_LEFT_DOWN, self.show_keypad)
        vbox.Add(self.max_temp_input, flag=wx.LEFT | wx.EXPAND, border=10)

        start_button = wx.Button(panel, label="Start")
        stop_button = wx.Button(panel, label="Stop")
        vbox.Add(start_button, flag=wx.LEFT | wx.TOP, border=10)
        vbox.Add(stop_button, flag=wx.LEFT | wx.TOP, border=10)

        panel.SetSizer(vbox)

        start_button.Bind(wx.EVT_BUTTON, self.on_start)
        stop_button.Bind(wx.EVT_BUTTON, self.on_stop)
        #self.GetParent().heater_status_label.SetLabel("OFF")
        #self.GetParent().heater_status_label.SetForegroundColour(COLOR_RED)
        ok_button = wx.Button(panel, label="OK")
        vbox.Add(ok_button, flag=wx.ALIGN_CENTER | wx.ALL, border=10)
        ok_button.Bind(wx.EVT_BUTTON, self.on_ok)

    def show_keypad(self, event):
        text_ctrl = event.GetEventObject()
        if text_ctrl is None:
            return
        keypad = NumericKeypad(self)
        try:
            if keypad.ShowModal() == wx.ID_OK:
                value = keypad.get_value()
                if text_ctrl is not None:
                    text_ctrl.SetValue(value)
        except Exception as e:
            print(f"Error in keypad: {e}")
        keypad.Destroy()

    def on_start(self, event):
        sv_value = self.temp_input.GetValue()
        max_temp_value = self.max_temp_input.GetValue()
    # Update settings and GUI
        self.settings["heater_1"]["sv"] = sv_value
        self.settings["heater_1"]["max_temp"] = max_temp_value
        self.GetParent().sv_label.SetLabel(f"SV: {sv_value}")
        self.GetParent().heater_status_label.SetLabel("ON")
        self.GetParent().heater_status_label.SetForegroundColour(COLOR_BLUE)
        self.GetParent().save_settings()
    # Relay ON via worker/queue (no direct send_pulse() call)
        #self.GetParent().relay_controller.send_pulse(4)
        self.GetParent().button_command_queue.put({"cmd": "relay_pulse", "channel": 4, "duration": 1.0})
    # Set SV and start TK4 via worker/queue
        self.GetParent().button_command_queue.put({"cmd": "set_tk4_sv", "address": 2, "value": float(sv_value)})
        self.GetParent().button_command_queue.put({"cmd": "start_tk4", "address": 2})
        self.EndModal(wx.ID_OK)


    def on_stop(self, event):
    # Stop TK4 via worker/queue
        self.GetParent().button_command_queue.put({"cmd": "stop_tk4", "address": 2})
    # Relay OFF via worker/queue
        #self.GetParent().relay_controller.send_pulse(5)
        self.GetParent().button_command_queue.put({"cmd": "relay_pulse", "channel": 5, "duration": 1.0})
        self.GetParent().heater_status_label.SetLabel("OFF")
        self.GetParent().heater_status_label.SetForegroundColour(COLOR_RED)
        self.EndModal(wx.ID_OK)

    def on_ok(self, event):
        sv_value = self.temp_input.GetValue()
        max_temp_value = self.max_temp_input.GetValue()
        self.settings["heater_1"]["sv"] = sv_value
        self.settings["heater_1"]["max_temp"] = max_temp_value
        self.GetParent().sv_label.SetLabel(f"SV: {sv_value}")
        #self.GetParent().heater_status_label.SetLabel("ON")
        #self.GetParent().heater_status_label.SetForegroundColour(COLOR_BLUE)
        self.GetParent().save_settings()
    # Send SV to TK4 via queue (same as Start)
        self.GetParent().button_command_queue.put({"cmd": "set_tk4_sv", "address": 2, "value": float(sv_value)})
        #self.GetParent().button_command_queue.put({"cmd": "start_tk4", "address": 2})
        #self.GetParent().button_command_queue.put({"cmd": "relay_pulse", "channel": 4, "duration": 1.0})
        self.EndModal(wx.ID_OK)



class Heater2Dialog(wx.Dialog):
    def __init__(self, parent, coil_sv_label, coil_pv_label, reactor_temp_label, heater_status_label, settings):
        super().__init__(parent, title="Heater Settings", size=(300, 300))
        self.coil_sv_label = coil_sv_label
        self.coil_pv_label = coil_pv_label
        self.reactor_temp_label = reactor_temp_label
        self.settings = settings
        
        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)

        # Add widgets for setting coil SV value
        vbox.Add(wx.StaticText(panel, label="Set Temperature (SV):"), flag=wx.LEFT | wx.TOP, border=10)
        self.coil_sv_input = wx.TextCtrl(panel, value=str(self.settings["heater_2"]["coil_sv"]))
        self.coil_sv_input.Bind(wx.EVT_LEFT_DOWN, self.show_keypad)
        vbox.Add(self.coil_sv_input, flag=wx.LEFT | wx.EXPAND, border=10)

        # Add widgets for setting maximum coil temperature
        vbox.Add(wx.StaticText(panel, label="Set Maximum Temperature:"), flag=wx.LEFT | wx.TOP, border=10)
        self.coil_max_temp_input = wx.TextCtrl(panel, value=str(self.settings["heater_2"]["coil_max_temp"]))
        self.coil_max_temp_input.Bind(wx.EVT_LEFT_DOWN, self.show_keypad)
        vbox.Add(self.coil_max_temp_input, flag=wx.LEFT | wx.EXPAND, border=10)

        # Add widgets for setting maximum reactor temperature
        vbox.Add(wx.StaticText(panel, label="Set Maximum Reactor Temperature:"), flag=wx.LEFT | wx.TOP, border=10)
        self.reactor_max_temp_input = wx.TextCtrl(panel, value=str(self.settings["heater_2"]["reactor_max_temp"]))
        self.reactor_max_temp_input.Bind(wx.EVT_LEFT_DOWN, self.show_keypad)
        vbox.Add(self.reactor_max_temp_input, flag=wx.LEFT | wx.EXPAND, border=10)

        # Add Start and Stop buttons
        start_button = wx.Button(panel, label="Start Heater")
        stop_button = wx.Button(panel, label="Stop Heater")
        vbox.Add(start_button, flag=wx.LEFT | wx.TOP, border=10)
        vbox.Add(stop_button, flag=wx.LEFT | wx.TOP, border=10)

        panel.SetSizer(vbox)

        # Bind button events
        start_button.Bind(wx.EVT_BUTTON, self.on_start)
        stop_button.Bind(wx.EVT_BUTTON, self.on_stop)
        #self.GetParent().heater_2_status_label.SetLabel("OFF")
        #self.GetParent().heater_2_status_label.SetForegroundColour(COLOR_RED)
        ok_button = wx.Button(panel, label="OK")
        vbox.Add(ok_button, flag=wx.ALIGN_CENTER | wx.ALL, border=10)
        ok_button.Bind(wx.EVT_BUTTON, self.on_ok)

    def show_keypad(self, event):
        text_ctrl = event.GetEventObject()
        if text_ctrl is None:
            return
        keypad = NumericKeypad(self)
        try:
            if keypad.ShowModal() == wx.ID_OK:
                value = keypad.get_value()
                if text_ctrl is not None:
                    text_ctrl.SetValue(value)
        except Exception as e:
            print(f"Error in keypad: {e}")
        keypad.Destroy()

    def on_start(self, event):
        coil_sv_value = self.coil_sv_input.GetValue()
        coil_max_temp_value = self.coil_max_temp_input.GetValue()
        reactor_max_temp_value = self.reactor_max_temp_input.GetValue()

    # Update GUI and settings
        if coil_sv_value:
            self.coil_sv_label.SetLabel(f"SV: {coil_sv_value}")
            self.settings["heater_2"]["coil_sv"] = coil_sv_value
        if coil_max_temp_value:
            self.settings["heater_2"]["coil_max_temp"] = coil_max_temp_value
        if reactor_max_temp_value:
            self.settings["heater_2"]["reactor_max_temp"] = reactor_max_temp_value

        #self.settings["heater_2"]["status"] = "ON"
        self.GetParent().heater_2_status_label.SetLabel("ON")
        self.GetParent().heater_2_status_label.SetForegroundColour(COLOR_BLUE)
        self.GetParent().save_settings()
        #self.GetParent().relay_controller.send_pulse(2)  # Relay 1 ON for 1s
        self.GetParent().button_command_queue.put({"cmd": "set_tk4_sv", "address": 1, "value": float(coil_sv_value)})
        self.GetParent().button_command_queue.put({"cmd": "start_tk4", "address": 1})
        self.GetParent().button_command_queue.put({"cmd": "relay_pulse", "channel": 2, "duration": 1.0})
        self.EndModal(wx.ID_OK)
        
    def on_stop(self, event):
        #self.GetParent().relay_controller.send_pulse(3)
        self.GetParent().button_command_queue.put({"cmd": "stop_tk4", "address": 1})
        self.GetParent().button_command_queue.put({"cmd": "relay_pulse", "channel": 3, "duration": 1.0})
        self.GetParent().heater_2_status_label.SetLabel("OFF")
        self.GetParent().heater_2_status_label.SetForegroundColour(COLOR_RED)

        self.EndModal(wx.ID_OK)

    def on_ok(self, event):
        coil_sv_value = self.coil_sv_input.GetValue()
        coil_max_temp_value = self.coil_max_temp_input.GetValue()
        reactor_max_temp_value = self.reactor_max_temp_input.GetValue()
    # Update settings and GUI
        if coil_sv_value:
            self.settings["heater_2"]["coil_sv"] = coil_sv_value
            self.coil_sv_label.SetLabel(f"SV: {coil_sv_value}")
        if coil_max_temp_value:
            self.settings["heater_2"]["coil_max_temp"] = coil_max_temp_value
        if reactor_max_temp_value:
            self.settings["heater_2"]["reactor_max_temp"] = reactor_max_temp_value
        self.GetParent().save_settings()
    # Only set SV in TK4, do NOT start heater
        self.GetParent().button_command_queue.put({
            "cmd": "set_tk4_sv",
            "address": 1,
            "value": float(coil_sv_value)
        })
        self.EndModal(wx.ID_OK)



class MFCSettingsDialog(wx.Dialog):
    def __init__(self, parent, sv_labels, status_labels, settings):
        super().__init__(parent, title="MFC Settings", size=(400, 450))
        self.sv_labels = sv_labels
        self.status_labels = status_labels
        self.settings = settings

        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)

        # Add input fields for each channel
        self.inputs = {}
        for i, channel in enumerate(["CH4", "O2", "N2", "H2"]):
            hbox = wx.BoxSizer(wx.HORIZONTAL)
            hbox.Add(wx.StaticText(panel, label=f"{channel} SV:"), flag=wx.RIGHT, border=10)
            self.inputs[channel] = wx.TextCtrl(panel, value=f"{self.settings['mfc_setpoints'][i]:.2f}")
            self.inputs[channel].Bind(wx.EVT_LEFT_DOWN, self.show_keypad)
            hbox.Add(self.inputs[channel], proportion=1)
            vbox.Add(hbox, flag=wx.EXPAND | wx.ALL, border=10)

        # Add Start and Stop buttons for each channel
        for i, channel in enumerate(["CH4", "O2", "N2", "H2"]):
            hbox = wx.BoxSizer(wx.HORIZONTAL)
            start_button = wx.Button(panel, label=f"Start {channel}")
            stop_button = wx.Button(panel, label=f"Stop {channel}")
            hbox.Add(start_button, flag=wx.RIGHT, border=10)
            hbox.Add(stop_button, flag=wx.RIGHT, border=10)
            vbox.Add(hbox, flag=wx.EXPAND | wx.ALL, border=10)

            # Bind button events
            start_button.Bind(wx.EVT_BUTTON, lambda event, ch=channel: self.on_start(event, ch))
            stop_button.Bind(wx.EVT_BUTTON, lambda event, ch=channel: self.on_stop(event, ch))

        # Add OK button
        ok_button = wx.Button(panel, label="OK")
        vbox.Add(ok_button, flag=wx.ALIGN_CENTER | wx.ALL, border=10)
        ok_button.Bind(wx.EVT_BUTTON, self.on_ok)

        panel.SetSizer(vbox)

    def show_keypad(self, event):
        text_ctrl = event.GetEventObject()
        if text_ctrl is None:
            return
        keypad = NumericKeypad(self)
        try:
            if keypad.ShowModal() == wx.ID_OK:
                value = keypad.get_value()
                if text_ctrl is not None:
                    text_ctrl.SetValue(value)
        except Exception as e:
            print(f"Error in keypad: {e}")
        keypad.Destroy()

    def on_start(self, event, channel):
        """Handle Start button click for a channel."""
        sv_value = self.inputs[channel].GetValue()
        try:
            sv_float = float(sv_value)
        except Exception:
            wx.MessageBox("Please enter a valid number for SV.", "Input Error", wx.ICON_WARNING)
            return

    # Update the SV label in the main GUI, always two decimals, always blue
        self.sv_labels[channel].SetLabel(f"{sv_float:.2f}")
        self.sv_labels[channel].SetForegroundColour(wx.Colour(50, 120, 220))  # COLOR_BLUE

    # Update the status label to ON and green
        self.status_labels[channel].SetLabel("ON")
        self.status_labels[channel].SetForegroundColour(wx.Colour(50, 180, 50))  # COLOR_GREEN

    # Update settings
        self.settings["mfc_setpoints"][self._channel_index(channel)] = sv_float

    # Send commands to worker thread
        self.GetParent().button_command_queue.put({
            "cmd": "set_mfc_flow",
            "channel": self._channel_index(channel) + 1,
            "value": sv_float
        })
        self.GetParent().button_command_queue.put({
            "cmd": "on_off_mfc",
            "channel": self._channel_index(channel) + 1,
            "state": True
        })


    def on_stop(self, event, channel):
        """Handle Stop button click for a channel."""
        self.status_labels[channel].SetLabel("OFF")
        self.status_labels[channel].SetForegroundColour(COLOR_RED)
        self.GetParent().button_command_queue.put({
            "cmd": "on_off_mfc",
            "channel": self._channel_index(channel) + 1,
            "state": False
        })

    def on_ok(self, event):
        for channel in ["CH4", "O2", "N2", "H2"]:
            sv_value = self.inputs[channel].GetValue()
            try:
                sv_float = float(sv_value)
            except Exception:
                wx.MessageBox(f"Please enter a valid number for {channel} SV.", "Input Error", wx.ICON_WARNING)
                return
            self.settings["mfc_setpoints"][self._channel_index(channel)] = sv_float
            self.sv_labels[channel].SetLabel(f"{sv_float:.2f}")
        # Send SV to the device (NEW)
            self.GetParent().button_command_queue.put({
                "cmd": "set_mfc_flow",
                "channel": self._channel_index(channel) + 1,
                "value": sv_float
            })
        self.GetParent().save_settings()
        self.EndModal(wx.ID_OK)



    def _channel_index(self, ch):
        return ["CH4", "O2", "N2", "H2"].index(ch)


class PressureDialog(wx.Dialog):
    def __init__(self, parent, sensor_id, settings):
        super().__init__(parent, title=f"Sensor {sensor_id} Settings", size=(300, 250))
        self.sensor_id = sensor_id
        self.settings = settings

        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)

        # Add widgets for setting maximum pressure
        vbox.Add(wx.StaticText(panel, label=f"Set Maximum Pressure for Sensor {sensor_id} (bar):"), flag=wx.LEFT | wx.TOP, border=10)
        self.max_pressure_input = wx.TextCtrl(panel, value=str(self.settings["sensor_settings"][f"sensor_{sensor_id}_max_pressure"]))
        self.max_pressure_input.Bind(wx.EVT_LEFT_DOWN, self.show_keypad)
        vbox.Add(self.max_pressure_input, flag=wx.LEFT | wx.EXPAND, border=10)

        # Add OK and Cancel buttons
        ok_button = wx.Button(panel, label="OK")
        cancel_button = wx.Button(panel, label="Cancel")
        vbox.Add(ok_button, flag=wx.LEFT | wx.TOP, border=10)
        vbox.Add(cancel_button, flag=wx.LEFT | wx.TOP, border=10)

        panel.SetSizer(vbox)

        # Bind button events
        ok_button.Bind(wx.EVT_BUTTON, self.on_ok)
        cancel_button.Bind(wx.EVT_BUTTON, self.on_cancel)

    def show_keypad(self, event):
        text_ctrl = event.GetEventObject()
        if text_ctrl is None:
            return
        keypad = NumericKeypad(self)
        try:
            if keypad.ShowModal() == wx.ID_OK:
                value = keypad.get_value()
                if text_ctrl is not None:
                    text_ctrl.SetValue(value)
        except Exception as e:
            print(f"Error in keypad: {e}")
        keypad.Destroy()
    def on_ok(self, event):
        max_pressure_value = self.max_pressure_input.GetValue()
        if max_pressure_value:
            self.settings["sensor_settings"][f"sensor_{self.sensor_id}_max_pressure"] = float(max_pressure_value)
            self.GetParent().save_settings()
        self.EndModal(wx.ID_OK)
    def on_cancel(self, event):
        """Close the dialog without saving."""
        self.EndModal(wx.ID_CANCEL)
    
class GasAnalyzer:
    def __init__(self, port="COM4"):
        self.port = port
        self.ser = None

    def connect(self):
        try:
            if self.ser is None or not self.ser.is_open:
                print(f"Connecting to gas analyzer on port {self.port}")
                self.ser = serial.Serial(self.port, baudrate=9600, timeout=1)
            return True
        except Exception as e:
            print(f"Gas analyzer connection error: {e}")
            self.ser = None
            return False

    def on_ok(self, event):
        max_pressure_value = self.max_pressure_input.GetValue()
        if max_pressure_value:
            self.settings["sensor_settings"][f"sensor_{self.sensor_id}_max_pressure"] = float(max_pressure_value)
            self.GetParent().save_settings()
        self.EndModal(wx.ID_OK)

    
    def close(self):
        if self.ser:
            self.ser.close()
            self.ser = None

    def read_gases(self):
        try:
            if self.ser is None or not self.ser.is_open:
                self.ser = serial.Serial(self.port, baudrate=9600, timeout=1)

            self.ser.reset_input_buffer()
            self.ser.write(bytes([0x11, 0x01, 0x01, 0xED]))

            time.sleep(0.2)

            data = self.ser.read(32)

            if len(data) >= 23:
                raw = data[3:23]
            elif len(data) == 20:
                raw = data
            else:
                print(f"Unexpected data length: {len(data)}")
                return None

            gas_names = ['CO', 'CO2', 'CH4', 'CnHm', 'H2', 'O2', 'C2H2', 'C2H4', 'HHV', 'N2']
            readings = {}

            for i, name in enumerate(gas_names):
                value_bytes = raw[i*2:i*2+2]
                value = int.from_bytes(value_bytes, byteorder='big', signed=False)
                readings[name] = value / 100.0

            return readings

        except Exception as e:
            print(f"Gas analyzer read error: {e}")
        # Close bad port so reconnect works next time
            if self.ser:
                try:
                    self.ser.close()
                except Exception:
                    pass
            self.ser = None
        return None






    def show_keypad(self, event):
        text_ctrl = event.GetEventObject()
        if text_ctrl is None:
            return
        keypad = NumericKeypad(self)
        try:
            if keypad.ShowModal() == wx.ID_OK:
                value = keypad.get_value()
                if text_ctrl is not None:
                    text_ctrl.SetValue(value)
        except Exception as e:
            print(f"Error in keypad: {e}")
        keypad.Destroy()

    def on_ok(self, event):
        """Save the maximum pressure setting."""
        max_pressure_value = self.max_pressure_input.GetValue()
        if max_pressure_value:
            self.settings["sensor_settings"][f"sensor_{self.sensor_id}_max_pressure"] = float(max_pressure_value)
        self.GetParent().save_settings()
        self.EndModal(wx.ID_OK)

    def on_cancel(self, event):
        """Close the dialog without saving."""
        self.EndModal(wx.ID_CANCEL)

class PowerMeterDialog(wx.Dialog):
    def __init__(self, parent, power_label, energy_label, integration_status_label):
        super().__init__(parent, title="Power Meter Settings", size=(450, 250))

        self.parent = parent  # needed to save settings
        self.settings = parent.settings
        self.power_label = power_label
        self.energy_label = energy_label
        self.integration_status_label = integration_status_label

        panel = wx.Panel(self)

        # Buttons for integration control
        start_button = wx.Button(panel, label="Start Integration", pos=(10, 10))
        stop_button = wx.Button(panel, label="Stop Integration", pos=(150, 10))
        reset_button = wx.Button(panel, label="Reset Integrated Value", pos=(290, 10))

        start_button.Bind(wx.EVT_BUTTON, self.start_integration)
        stop_button.Bind(wx.EVT_BUTTON, self.stop_integration)
        reset_button.Bind(wx.EVT_BUTTON, self.reset_integration)

        # Current scaling factor from settings
        wx.StaticText(panel, label="Current Scaling Factor:", pos=(10, 60))
        initial_value = str(self.settings.get("power_meter_scaling_factor", 10.0))
        self.scaling_factor_input = wx.TextCtrl(panel, value=initial_value, pos=(150, 60), size=(100, -1))
        self.scaling_factor_input.Bind(wx.EVT_LEFT_DOWN, self.show_keypad)

        # OK button to close the dialog
        ok_button = wx.Button(panel, label="OK", pos=(150, 120))
        ok_button.Bind(wx.EVT_BUTTON, self.on_ok)

    def start_integration(self, event):
        self.integration_status_label.SetLabel("ON")
        self.integration_status_label.SetForegroundColour(COLOR_BLUE)
        self.parent.device_manager.start_integration()

    def stop_integration(self, event):
        self.integration_status_label.SetLabel("OFF")
        self.integration_status_label.SetForegroundColour(COLOR_RED)
        self.parent.device_manager.stop_integration()

    def reset_integration(self, event):
        self.energy_label.SetLabel("0.0")
        self.parent.device_manager.reset_integration()

    def show_keypad(self, event):
        text_ctrl = event.GetEventObject()
        if text_ctrl is None:
            return
        keypad = NumericKeypad(self)
        try:
            if keypad.ShowModal() == wx.ID_OK:
                value = keypad.get_value()
                if text_ctrl is not None:
                    text_ctrl.SetValue(value)
        except Exception as e:
            print(f"Error in keypad: {e}")
        keypad.Destroy()

    def on_ok(self, event):
        try:
            value = float(self.scaling_factor_input.GetValue())
            self.settings["power_meter_scaling_factor"] = value
            self.parent.save_settings()
            print(f"Scaling factor set to: {value}")
        except ValueError:
            wx.MessageBox("Please enter a valid number for scaling factor.", "Input Error", wx.ICON_WARNING)
            return
        self.Close()



class CustomStatusBarRedirector:
    def __init__(self, status_label):
        self.status_label = status_label

    def write(self, message):
        """Write messages to the custom status bar."""
        if message.strip():  # Ignore empty messages
            self.status_label.SetLabel(message.strip())

    def flush(self):
        """Flush method for compatibility (does nothing)."""
        pass


class DeviceManager:
    def __init__(self, settings, lock=None):
        self.settings = settings
        self.lock = lock or threading.Lock()
        self.power_port = settings.get("PM_PORT", "COM3")
        #self.pm = PowerMeter(port=self.power_port)
        self.gas_port = settings.get("GAS_ANALYZER_PORT", "COM4")
        self.gas_analyzer = GasAnalyzer(self.gas_port)
        self.rs485_port = settings.get("RS485_PORT", "COM5")
        self.client = ModbusSerialClient(
            port=self.rs485_port, baudrate=9600, stopbits=2, bytesize=8,
            parity='N', timeout=0.2, retries=0
        )
        self.client.connect()
        self.psm4 = PSM4Controller(self.client)
        self.mfc = MFCController(port=self.rs485_port)
        self.mfm = MFMFlowMeter(port=self.rs485_port)
        self.configure_power_meter()
        
    def read_gas_analyzer(self):
        """Read all gas values as a dict, or None if not connected."""
        try:
            return self.gas_analyzer.read_gases()
        except Exception as e:
            print(f"Gas analyzer read error: {e}")
            return None

    def read_temperature(self, slave_id):
        try:
            with self.lock:
                result = self.client.read_input_registers(0x03E8, count=2, slave=slave_id)
            if result.isError():
                return None
            raw, decimal = result.registers
            if raw == 31000:
                return None
            return raw / (10 ** decimal)
        except Exception as e:
            print(f"TK4_{slave_id}: not connected")
            return None


    def set_sv(self, slave_id, temperature):
        try:
            with self.lock:
                return not self.client.write_register(0x0000, int(temperature), slave=slave_id).isError()
        except Exception as e:
            print(f"[MODBUS] Failed to set SV on ID {slave_id}: {e}")
            return False

    def start_heater(self, slave_id):
        try:
            with self.lock:
                return not self.client.write_register(0x0032, 0, slave=slave_id).isError()
        except Exception as e:
            print(f"[MODBUS] Failed to start heater ID {slave_id}: {e}")
            return False

    def stop_heater(self, slave_id):
        try:
            with self.lock:
                return not self.client.write_register(0x0032, 1, slave=slave_id).isError()
        except Exception as e:
            print(f"[MODBUS] Failed to stop heater ID {slave_id}: {e}")
            return False

    def read_power_meter(self):
        try:
            with serial.Serial(self.power_port, baudrate=9600, timeout=2) as ser:
            # Read power
                ser.write(b":NUMERIC:NORMAL:VALUE?3\r\n")
                time.sleep(0.2)
                power_line = ser.readline().decode().strip()
                power = float(power_line) if power_line else None

            # Read energy
                ser.write(b":NUMERIC:NORMAL:VALUE?4\r\n")
                time.sleep(0.2)
                energy_line = ser.readline().decode().strip()
                energy = float(energy_line) if energy_line else None

                return power, energy
        except Exception as e:
            print("Power Meter: not connected")
            return None, None

    def configure_power_meter(self):
        try:
            with serial.Serial(self.power_port, baudrate=9600, timeout=2) as ser:
                ser.write(b":COMMunicate:REMote ON\r\n")
                time.sleep(0.2)
                ser.write(b":NUMERIC:NORMAL:ITEM4 WH,1\r\n")
                time.sleep(0.2)
                ser.write(b":INTEGrate:MODE MANUAL\r\n")
                time.sleep(0.2)
                ser.write(b":INTEGrate:FUNCtion WP\r\n")
                time.sleep(0.5)
        except Exception as e:
            print(f"Power meter configuration error: {e}")

    
    def start_integration(self):
        try:
            with serial.Serial(self.power_port, baudrate=9600, timeout=2) as ser:
                ser.write(b":INTEGrate:RESet\r\n")
                time.sleep(0.5)
                ser.write(b":INTEGrate:STARt\r\n")
        except Exception as e:
            print(f"Power meter start integration error: {e}")


    def stop_integration(self):
        try:
            with serial.Serial(self.power_port, baudrate=9600, timeout=0.2) as ser:
                ser.write(b":INTEGrate:STOP\r\n")
        except Exception as e:
         print(f"Power meter stop integration error: {e}")

    def reset_integration(self):
        try:
            with serial.Serial(self.power_port, baudrate=9600, timeout=0.2) as ser:
                ser.write(b":INTEGrate:RESet\r\n")
                
        except Exception as e:
            print(f"Power meter reset integration error: {e}")



    def set_mfc_flow(self, channel, value):
        try:
            #with self.lock:
            return self.mfc.set_flow(channel, value)
        except Exception as e:
            print(f"MFC set flow error: {e}")
            return False

    def on_off_mfc(self, channel, state):
        try:
            
            return self.mfc.on_off(channel, state)
        except Exception as e:
            print(f"MFC on/off error: {e}")
            return False

    def read_mfc_flow(self, channel):
        """Read the PV (process value/flow) for a single MFC channel."""
        try:
            return self.mfc.read_flow(channel)
        except Exception as e:
            print(f"MFC read flow error: {e}")
            return None

    def read_all_mfc_flows(self):
        """Read all MFC flows (channels 1-4). Returns a list of 4 floats or None if not connected."""
        try:
            return self.mfc.read_all_flows()
        except Exception as e:
            print(f"MFC read all flows error: {e}")
            return [None] * 4


    def read_mfm_flow(self):
        """
    Read the MFM flow value.
    Returns a float or None if not connected.
    """
        try:
            return self.mfm.read_flow()
        except Exception as e:
            print(f"MFM read flow error: {e}")
            return None


    def read_pressures(self):
        try:
            with self.lock:
                return self.psm4.read_pressures()
        except Exception as e:
            print(f"PSM4 read error: {e}")
            return ["NC"] * 4


class PSM4Controller:
    """
    Modbus RTU controller for PSM4 pressure sensor (4 channels).
    Assumes slave address 7 and standard register layout.
    """
    def __init__(self, client):
        self.client = client  # pymodbus ModbusSerialClient
        self.lock = threading.Lock()

    def read_pressures(self):
        """
        Read all 4 pressure channels (slave ID 7).
        Returns a list of 4 values (floats in bar, or "NC" if not connected/invalid).
        """
        pressures = []
        # Base register addresses for each channel (from your v1.0.4 code)
        base_addrs = [0x03E8, 0x03ED, 0x03F2, 0x03F7]
        try:
            with self.lock:
                for addr in base_addrs:
                    response = self.client.read_input_registers(
                        address=addr + 1,
                        count=2,
                        slave=7
                    )
                    if response.isError():
                        pressures.append("NC")
                        continue
                    raw, decimal = response.registers
                    try:
                        val = raw / (10 ** decimal) / 10.0  # per your scaling
                    except Exception:
                        pressures.append("NC")
                        continue
                    # If value > 20 bar, show NC (sensor disconnected or error)
                    if val > 20.0:
                        pressures.append("NC")
                    else:
                        pressures.append(round(val, 2))
            return pressures
        except Exception as e:
            #print(f"PSM4 read error: {str(e)}")
            return ["NC"] * 4

    def close(self):
        if self.client:
            self.client.close()

class MFCController:
    def __init__(self, port="COM6", channels=[1,2,3,4]):
        self.port = port
        self.channels = channels

    def _checksum(self, data):
        cs = 0
        for b in data.encode():
            cs ^= b
        return f"{cs:02X}"

    def send_command(self, channel, cmd, addr, value=""):
        frame_wo_cs = f":{channel:02d}{cmd}{addr}{value}"
        cs = self._checksum(frame_wo_cs)
        frame = f"{frame_wo_cs}{cs}\r".encode()
        with serial.Serial(
            port=self.port, baudrate=9600, bytesize=8,
            parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout=0.1
        ) as ser:
            ser.write(frame)
            resp = ser.read_until(b'\r')
        return resp

    def read_flow(self, channel):
        try:
            resp = self.send_command(channel, "03", "0038")
            if not resp or not resp.startswith(b':'):
                return None
            data_ascii = resp[7:15].decode()
            return struct.unpack('>f', bytes.fromhex(data_ascii))[0]
        except Exception as e:
            print("MFC: not connected")
            return None

    
    def set_flow(self, channel, value):
        hexval = struct.pack('!f', float(value)).hex().upper()
        resp = self.send_command(channel, "01", "07", hexval)
        return resp.startswith(f":{channel:02d}81".encode())

    def on_off(self, channel, state):
        val = "01" if state else "00"
        resp = self.send_command(channel, "58", "02", val)
        return resp.startswith(f":{channel:02d}D8".encode())

    def read_all_flows(self):
        return [self.read_flow(ch) for ch in self.channels]

class MFMFlowMeter:
    def __init__(self, port='COM6', device_id=1):
        self.port = port
        self.device_id = device_id

    def _checksum(self, frame_wo_cs):
        cs = 0
        for b in frame_wo_cs.encode():
            cs ^= b
        return f"{cs:02X}"

    def build_read_flow_frame(self):
        id_ascii = f"{self.device_id:02X}"
        frame_wo_cs = f":{id_ascii}0300"
        cs = self._checksum(frame_wo_cs)
        return f"{frame_wo_cs}{cs}\r".encode()

    def read_flow(self):
        with serial.Serial(
            port=self.port, baudrate=9600, bytesize=8,
            parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout=0.1
        ) as ser:
            frame = self.build_read_flow_frame()
            ser.write(frame)
            resp = ser.read_until(b'\r')
            if not resp.startswith(b':'):
                return None
            if len(resp) < 16:
                return None
            try:
                data_ascii = resp[7:15].decode('ascii')
                data_bytes = bytes.fromhex(data_ascii)
                return struct.unpack('>f', data_bytes)[0]
            except Exception:
                return None
    
    


class ControlGUI(wx.Frame):
    def __init__(self, parent, title):
        super().__init__(parent, title=title, size=(1280, 800), style=wx.DEFAULT_FRAME_STYLE & ~wx.RESIZE_BORDER)
        # Initialize latch dictionaries FIRST
        self.overtemp_latched = {1: False, 2: False, 3: False}  # 1: Heater2, 2: Heater1, 3: Reactor
        self.overpress_latched = {1: False, 2: False, 3: False}
        #log_filename = f"process_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.logger = Logger()
        self.polling_paused = False
        self.device_status = {
            "TK4_1": False,
            "TK4_2": False,
            "TK4_3": False,
            "TK4_4": False,
            "TK4_5": False,
            "TK4_6": False,
            "PSM4": False,
            "MFC": False,
            "MFM": False,
            "PowerMeter": False,
            "GasAnalyzer": False
        }
        self.button_command_queue = queue.Queue()
        self.settings = self.load_settings()
        self.modbus_lock = threading.Lock()
        self.running = True
        self.worker_thread = threading.Thread(target=self.button_command_handler, daemon=True)
        self.worker_thread.start()

        self.device_manager = DeviceManager(self.settings, lock=self.modbus_lock)
        self.relay_controller = ModbusRelayController(self.device_manager.client, lock=self.modbus_lock, slave_id=8)
        self.tk4_ids = [1, 2, 3, 4, 5, 6]  # 1: Heater 2, 2: Heater 1, 3: Reactor, 4-6: extra sensors
        
        self.heater_1_id = 2
        self.heater_2_id = 1
        self.sensor_ids = [3, 4, 5]
        #self.relay_controller = ModbusRelayController(self.device_manager.client, lock=self.modbus_lock, slave_id=8)

        self.settings = self.load_settings()
        self.setup_gui()
        self.start_datetime_updates()
        self.start_device_updates()

        self.Bind(wx.EVT_CLOSE, self.on_close)
        self.alarm_active = False
        self.check_emergency_conditions()
        


    # --- GUI SETUP ---
    def setup_gui(self):
        SHARED_FONT = wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        COLOR_BLUE = wx.Colour(0, 0, 255)
        COLOR_RED = wx.Colour(255, 0, 0)
        COLOR_GREEN = wx.Colour(0, 255, 0)
        LARGE_FONT = wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        
        self.original_background_image = wx.Image("background.png", wx.BITMAP_TYPE_PNG)
        self.background_bitmap = self.original_background_image.ConvertToBitmap()
        self.background = wx.StaticBitmap(self, -1, self.background_bitmap)
        self.background.Bind(wx.EVT_LEFT_DOWN, self.on_mouse_click)

        # Status bar
        self.status_bar_panel = wx.Panel(self.background, pos=(0, 0), size=(1280, 50))
        self.status_bar_panel.SetBackgroundColour(wx.Colour(0, 0, 0))
        self.datetime_label = wx.StaticText(self.status_bar_panel, label="", pos=(10, 15))
        self.datetime_label.SetFont(SHARED_FONT)
        self.datetime_label.SetForegroundColour(wx.Colour(255, 255, 255))
        self.status_bar_label = wx.StaticText(self.status_bar_panel, label="Welcome to POX Control System!", pos=(200, 15))
        self.status_bar_label.SetFont(SHARED_FONT)
        self.status_bar_label.SetForegroundColour(wx.Colour(255, 255, 255))
        self.version_label = wx.StaticText(self.status_bar_panel, label="APGREEN v.2.0.0", pos=(1100, 15))
        self.version_label.SetFont(SHARED_FONT)
        self.version_label.SetForegroundColour(wx.Colour(255, 255, 255))
        self.selftest_btn = wx.Button(self.status_bar_panel, label="Selftest", pos=(900, 10), size=(80, 30))
        self.selftest_btn.Bind(wx.EVT_BUTTON, lambda evt: self.run_selftest())
        self.plot_btn = wx.Button(self.status_bar_panel, label="Plot", pos=(1000, 10), size=(80, 30))
        self.plot_btn.Bind(wx.EVT_BUTTON, self.on_plot)
        self.selftest_btn.SetFont(SHARED_FONT)
        self.plot_btn.SetFont(SHARED_FONT)
       
        #sys.stdout = CustomStatusBarRedirector(self.status_bar_label)

        # Heater 1 Panel
        self.text_box_outline = wx.Panel(self.background, pos=(365, 160), size=(100, 80), style=wx.SIMPLE_BORDER)
        self.text_box_outline.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.sv_label = wx.StaticText(self.text_box_outline, label=f"SV: {self.settings['heater_1'].get('sv', '--')}", pos=(10, 10))
        self.pv_label = wx.StaticText(self.text_box_outline, label="PV: --", pos=(10, 30))
        self.heater_status_label = wx.StaticText(self.text_box_outline, label="OFF", pos=(10, 50))
        for lbl, color in [(self.sv_label, COLOR_BLUE), (self.pv_label, COLOR_RED), (self.heater_status_label, COLOR_RED)]:
            lbl.SetFont(SHARED_FONT)
            lbl.SetForegroundColour(color)

        # Heater 2 Panel
        self.heater_2_panel = wx.Panel(self.background, pos=(810, 250), size=(120, 95), style=wx.SIMPLE_BORDER)
        self.heater_2_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.coil_sv_label = wx.StaticText(self.heater_2_panel, label=f"SV: {self.settings['heater_2']['coil_sv']}", pos=(10, 10))
        self.coil_pv_label = wx.StaticText(self.heater_2_panel, label=f"PV: {self.settings['heater_2']['coil_pv']}", pos=(10, 30))
        self.reactor_temp_label = wx.StaticText(self.heater_2_panel, label=f"Reactor: {self.settings['heater_2']['reactor_temp']}", pos=(10, 50))
        self.heater_2_status_label = wx.StaticText(
        self.heater_2_panel,
        label="OFF",  # Always initialize as OFF or whatever you want as default
        pos=(10, 70))

        for lbl, color in [
            (self.coil_sv_label, COLOR_BLUE), (self.coil_pv_label, COLOR_RED),
            (self.reactor_temp_label, COLOR_BLUE), (self.heater_2_status_label, COLOR_RED)
        ]:
            lbl.SetFont(SHARED_FONT)
            lbl.SetForegroundColour(color)
        
        # Reactor temperature label in Heater 2 panel (ID=3)
        #self.reactor_temp_label = wx.StaticText(self.heater_2_panel, label=f"Reactor: --", pos=(10, 50))
        self.reactor_temp_label.SetFont(SHARED_FONT)
        self.reactor_temp_label.SetForegroundColour(COLOR_BLUE)

# Read-only sensors (IDs 4, 5, 6)
        self.sensor_labels = []
        sensor_positions = [(130, 170), (620, 420), (620, 535)]  # Adjust as needed
        for i in range(3):
            panel = wx.Panel(self.background, pos=sensor_positions[i], size=(100, 40), style=wx.SIMPLE_BORDER)
            panel.SetBackgroundColour(wx.Colour(255, 255, 255))
            label = wx.StaticText(panel, label=f"Temp: --", pos=(10, 10))
            label.SetFont(SHARED_FONT)
            label.SetForegroundColour(COLOR_BLUE)
            self.sensor_labels.append(label)

        
        
        # Pressure Panels
        self.pressure_panel = wx.Panel(self.background, pos=(130, 90), size=(90, 40), style=wx.SIMPLE_BORDER)
        self.pressure_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.pressure_label = wx.StaticText(self.pressure_panel, label=self.settings["pressure_values"]["sensor_1"], pos=(10, 10))
        self.pressure_label.SetFont(SHARED_FONT)
        self.pressure_label.SetForegroundColour(COLOR_RED)
        self.pressure_2_panel = wx.Panel(self.background, pos=(800, 110), size=(90, 40), style=wx.SIMPLE_BORDER)
        self.pressure_2_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.pressure_2_label = wx.StaticText(self.pressure_2_panel, label="-- bar", pos=(10, 10))
        self.pressure_2_label.SetFont(SHARED_FONT)
        self.pressure_2_label.SetForegroundColour(COLOR_RED)
        self.pressure_3_panel = wx.Panel(self.background, pos=(620, 610), size=(90, 40), style=wx.SIMPLE_BORDER)
        self.pressure_3_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.pressure_3_label = wx.StaticText(self.pressure_3_panel, label="-- bar", pos=(10, 10))
        self.pressure_3_label.SetFont(SHARED_FONT)
        self.pressure_3_label.SetForegroundColour(COLOR_RED)

        

        # MFC Panel
        self.mfc_panel = wx.Panel(self.background, pos=(100, 315), size=(330, 130), style=wx.SIMPLE_BORDER)
        self.mfc_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        wx.StaticText(self.mfc_panel, label="Channel", pos=(10, 5)).SetFont(SHARED_FONT)
        wx.StaticText(self.mfc_panel, label="SV", pos=(100, 5)).SetFont(SHARED_FONT)
        wx.StaticText(self.mfc_panel, label="PV", pos=(180, 5)).SetFont(SHARED_FONT)
        wx.StaticText(self.mfc_panel, label="Status", pos=(260, 5)).SetFont(SHARED_FONT)
        self.mfc_panel.Bind(wx.EVT_LEFT_DOWN, self.open_mfc_settings)
        self.mfc_channels = ["CH4", "O2", "N2", "H2"]
        self.mfc_sv_labels = {}
        self.mfc_pv_labels = {}
        self.mfc_status_labels = {}
        for i, channel in enumerate(self.mfc_channels):
            y_offset = 25 + i * 25
            channel_label = wx.StaticText(self.mfc_panel, label=channel, pos=(10, y_offset))
            channel_label.SetFont(SHARED_FONT)
            self.mfc_sv_labels[channel] = wx.StaticText(self.mfc_panel, label=f"{self.settings['mfc_setpoints'][i]:.2f}", pos=(100, y_offset))

            self.mfc_sv_labels[channel].SetFont(SHARED_FONT)
            self.mfc_sv_labels[channel].SetForegroundColour(COLOR_BLUE)
            self.mfc_pv_labels[channel] = wx.StaticText(self.mfc_panel, label="--", pos=(180, y_offset))
            self.mfc_pv_labels[channel].SetFont(SHARED_FONT)
            self.mfc_pv_labels[channel].SetForegroundColour(COLOR_RED)
            status = "ON" if self.settings["mfc_states"][i] else "OFF"
            self.mfc_status_labels[channel] = wx.StaticText(self.mfc_panel, label=status, pos=(260, y_offset))
            self.mfc_status_labels[channel].SetFont(SHARED_FONT)
            self.mfc_status_labels[channel].SetForegroundColour(COLOR_BLUE if status == "ON" else COLOR_RED)

        # Gas Analyzer Panel
        self.gas_analyzer_panel = wx.Panel(self.background, pos=(980, 370), size=(280, 170), style=wx.SIMPLE_BORDER)
        self.gas_analyzer_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.gas_labels = {}
        gases_column_1 = ["CO", "CH4", "H2", "C2H2", "HHV"]
        gases_column_2 = ["CO2", "CnHm", "O2", "C2H4", "N2"]
        for i in range(5):
            gas_label_1 = wx.StaticText(self.gas_analyzer_panel, label=f"{gases_column_1[i]}:", pos=(10, 10 + i * 30))
            gas_label_1.SetFont(SHARED_FONT)
            self.gas_labels[gases_column_1[i]] = wx.StaticText(self.gas_analyzer_panel, label="-- %", pos=(70, 10 + i * 30))
            self.gas_labels[gases_column_1[i]].SetFont(SHARED_FONT)
            self.gas_labels[gases_column_1[i]].SetForegroundColour(COLOR_BLUE)
            gas_label_2 = wx.StaticText(self.gas_analyzer_panel, label=f"{gases_column_2[i]}:", pos=(150, 10 + i * 30))
            gas_label_2.SetFont(SHARED_FONT)
            self.gas_labels[gases_column_2[i]] = wx.StaticText(self.gas_analyzer_panel, label="-- %", pos=(210, 10 + i * 30))
            self.gas_labels[gases_column_2[i]].SetFont(SHARED_FONT)
            self.gas_labels[gases_column_2[i]].SetForegroundColour(COLOR_BLUE)

        # MFM Panel
        self.mfm_panel = wx.Panel(self.background, pos=(910, 660), size=(80, 40))
        self.mfm_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        self.mfm_label = wx.StaticText(self.mfm_panel, label="--", pos=(10, 10))
        self.mfm_label.SetFont(SHARED_FONT)
        self.mfm_label.SetForegroundColour(COLOR_RED)

        # Power Meter Panel
        self.power_meter_panel = wx.Panel(self.background, pos=(400, 630), size=(180, 150), style=wx.SIMPLE_BORDER)
        self.power_meter_panel.SetBackgroundColour(wx.Colour(255, 255, 255))
        wx.StaticText(self.power_meter_panel, label="Power (W):", pos=(10, 10)).SetFont(LARGE_FONT)
        self.power_label = wx.StaticText(self.power_meter_panel, label="--", pos=(120, 10))
        self.power_label.SetFont(LARGE_FONT)
        self.power_label.SetForegroundColour(COLOR_BLUE)
        wx.StaticText(self.power_meter_panel, label="Energy (Wh):", pos=(10, 40)).SetFont(LARGE_FONT)
        self.energy_label = wx.StaticText(self.power_meter_panel, label="--", pos=(120, 40))
        self.energy_label.SetFont(LARGE_FONT)
        self.energy_label.SetForegroundColour(COLOR_BLUE)
        wx.StaticText(self.power_meter_panel, label="Integration:", pos=(10, 70)).SetFont(LARGE_FONT)
        self.integration_status_label = wx.StaticText(self.power_meter_panel, label="OFF", pos=(120, 70))
        self.integration_status_label.SetFont(LARGE_FONT)
        self.integration_status_label.SetForegroundColour(COLOR_RED)
        self.power_meter_panel.Bind(wx.EVT_LEFT_DOWN, self.on_power_meter_click)
        wx.CallAfter(self.run_selftest)

    def on_plot(self, event):
    
        log_files = glob.glob(os.path.join(data_log_dir, "process_log_*.xlsx"))
        if not log_files:
            wx.MessageBox("No log files found in Data log folder.", "Plot Error", wx.ICON_ERROR)
            return
        latest_log = max(log_files, key=os.path.getmtime)
        dlg = PlotDialog(self, latest_log, data_log_dir)
        dlg.ShowModal()
        dlg.Destroy()



    def run_selftest(self):
        status = {}
    # TK4 controllers (IDs 1-6)
        for addr in [1, 2, 3, 4, 5, 6]:
            try:
                temp = self.device_manager.read_temperature(addr)
                print(f"Selftest: TK4_{addr} temp={temp}")
                status[f"TK4_{addr}"] = temp is not None
            except Exception:
                status[f"TK4_{addr}"] = False

    # PSM4 (pressure)
        try:
            pressures = self.device_manager.read_pressures()
            status["PSM4"] = any(isinstance(p, (float, int)) for p in pressures)
        except Exception:
            status["PSM4"] = False

    # MFC (try reading flow from channel 1)
        try:
            reply_queue = queue.Queue()
            self.button_command_queue.put({"cmd": "read_mfc", "channel": 1, "reply_queue": reply_queue})
            result = reply_queue.get(timeout=0.5)
            flow = result.get("value", None)
            status["MFC"] = flow is not None
        except Exception:
            status["MFC"] = False

    # MFM (use worker thread)
        try:
            reply_queue = queue.Queue()
            self.button_command_queue.put({"cmd": "read_mfm", "reply_queue": reply_queue})
            result = reply_queue.get(timeout=2)
            mfm_flow = result.get("value", None)
            status["MFM"] = mfm_flow is not None
        except Exception:
            status["MFM"] = False

    # Power Meter
        try:
            power, energy = self.device_manager.read_power_meter()
            status["PowerMeter"] = power is not None and energy is not None
        except Exception:
            status["PowerMeter"] = False

    # Gas Analyzer
        try:
            reply_queue = queue.Queue()
            self.button_command_queue.put({"cmd": "read_gas_analyzer", "reply_queue": reply_queue})
            result = reply_queue.get(timeout=2)
            gases = result.get("values", None)
            status["GasAnalyzer"] = gases is not None
        except Exception:
            status["GasAnalyzer"] = False



    # Update labels with self-test results
    # Store results
        self.device_status = status
        
    # Show popup with results
        #self.show_selftest_popup(status)
        available_ports = [port.device for port in serial.tools.list_ports.comports()]
        self.show_selftest_popup(status, available_ports)
        
    def show_selftest_popup(self, status, available_ports):
        msg = ""
        for dev, ok in status.items():
            msg += f"{dev}: {'OK' if ok else 'NOT CONNECTED'}\n"
        msg += "\nAvailable serial ports detected:\n"
        if available_ports:
            msg += "\n".join(available_ports)
        else:
            msg += "(No serial ports found)"
        msg += "\n\nIf a device is NOT CONNECTED, check your settings.json and available ports, then restart the program."
        dlg = wx.MessageDialog(self, msg, "Self-Test Results", wx.OK | wx.ICON_INFORMATION)
        dlg.ShowModal()
        dlg.Destroy()



    # --- DEVICE COMMAND HANDLER THREAD ---
    def button_command_handler(self):
        while self.running:
            try:
            # Use a timeout so the thread can check self.running and exit cleanly
                try:
                    cmd = self.button_command_queue.get(timeout=0.2)
                except queue.Empty:
                    continue

                cmd_type = cmd.get("cmd")
                reply_queue = cmd.get("reply_queue", None)

            # --- TK4 (Modbus) ---
                if cmd_type == "set_tk4_sv":
                    addr = cmd["address"]
                    val = cmd["value"]
                    try:
                        result = self.device_manager.set_sv(addr, val)
                    except Exception as e:
                        print(f"[Worker] set_tk4_sv error: {e}")
                        result = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "address": addr, "success": result})

                elif cmd_type == "start_tk4":
                    addr = cmd["address"]
                    try:
                        result = self.device_manager.start_heater(addr)
                    except Exception as e:
                        print(f"[Worker] start_tk4 error: {e}")
                        result = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "address": addr, "success": result})

                elif cmd_type == "stop_tk4":
                    addr = cmd["address"]
                    try:
                        result = self.device_manager.stop_heater(addr)
                    except Exception as e:
                        print(f"[Worker] stop_tk4 error: {e}")
                        result = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "address": addr, "success": result})

            # --- Relay (Modbus) ---
                elif cmd_type == "relay_pulse":
                    
                    channel = cmd["channel"]
                    duration = cmd.get("duration", 1.0)
                    print(f"Worker: Pulsing relay {channel} for {duration}s")
                    try:
                        self.relay_controller.send_pulse(channel, duration)
                        success = True
                    except Exception as e:
                        print(f"[Worker] relay_pulse error: {e}")
                        success = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "channel": channel, "success": success})

                elif cmd_type == "relay_open_all":
                    try:
                        self.relay_controller.open_all()
                        success = True
                    except Exception as e:
                        print(f"[Worker] relay_open_all error: {e}")
                        success = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "success": success})

                elif cmd_type == "relay_close_all":
                    try:
                        self.relay_controller.close_all()
                        success = True
                    except Exception as e:
                        print(f"[Worker] relay_close_all error: {e}")
                        success = False
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "success": success})

            # --- MFC/MFM (ASCII) - Requires port switching ---
                elif cmd_type in ("set_mfc_flow", "on_off_mfc", "read_mfc", "read_all_mfc", "read_mfm"):
                    try:
                    # Always close Modbus before ASCII
                        with self.modbus_lock:
                            if self.device_manager.client.is_socket_open():
                                self.device_manager.client.close()
                    # Perform the ASCII operation
                        if cmd_type == "set_mfc_flow":
                            ch = cmd["channel"]
                            val = cmd["value"]
                            result = self.device_manager.set_mfc_flow(ch, val)
                            if reply_queue:
                                reply_queue.put({"cmd": cmd_type, "channel": ch, "success": result})
                        elif cmd_type == "on_off_mfc":
                            ch = cmd["channel"]
                            state = cmd["state"]
                            result = self.device_manager.on_off_mfc(ch, state)
                            if reply_queue:
                                reply_queue.put({"cmd": cmd_type, "channel": ch, "success": result})
                        elif cmd_type == "read_mfc":
                            ch = cmd["channel"]
                            value = self.device_manager.read_mfc_flow(ch)
                            if reply_queue:
                                reply_queue.put({"cmd": cmd_type, "channel": ch, "value": value})
                        elif cmd_type == "read_all_mfc":
                            flows = self.device_manager.read_all_mfc_flows()
                            if reply_queue:
                                reply_queue.put({"cmd": cmd_type, "values": flows})
                        elif cmd_type == "read_mfm":
                            value = self.device_manager.read_mfm_flow()
                            if reply_queue:
                                reply_queue.put({"cmd": cmd_type, "value": value})
                    except Exception as e:
                        print(f"[Worker] {cmd_type} error: {e}")
                        if reply_queue:
                            reply_queue.put({"cmd": cmd_type, "success": False})
                    finally:
                    # Always reconnect Modbus after ASCII
                        with self.modbus_lock:
                            if not self.device_manager.client.is_socket_open():
                                self.device_manager.client.connect()
                                time.sleep(0.1)
                            self.relay_controller.client = self.device_manager.client
                            self.device_manager.psm4.client = self.device_manager.client

            # --- PSM4 (Modbus) ---
                elif cmd_type == "read_psm4":
                    try:
                        values = self.device_manager.read_pressures()
                    except Exception as e:
                        print(f"[Worker] read_psm4 error: {e}")
                        values = ["NC"] * 4
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "values": values})

            # --- Power Meter (if implemented) ---
                elif cmd_type == "read_power_meter":
                    try:
                        power, energy = self.device_manager.read_power_meter()
                    except Exception as e:
                        print(f"[Worker] read_power_meter error: {e}")
                        power, energy = None, None
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "power": power, "energy": energy})

            # --- Gas Analyzer (if implemented) ---
                elif cmd_type == "read_gas_analyzer":
                    try:
                        gases = self.device_manager.read_gas_analyzer()
                    except Exception as e:
                        print(f"[Worker] read_gas_analyzer error: {e}")
                        gases = None
                    if reply_queue:
                        reply_queue.put({"cmd": cmd_type, "values": gases})

                else:
                    print(f"[Worker] Unknown command: {cmd_type}")

            except Exception as e:
                print(f"Control error: {e}")

        print("button_command_handler thread exiting")


    def control_heater(self, address, state):
        try:
            if self.modbus_client is None or not getattr(self.modbus_client, 'connected', False):
                print(f"MODBUS client not initialized or not connected. Cannot control heater ID {address}.")
                return False
            value = 1 if state else 0
            response = self.modbus_client.write_register(address=0x0032, value=value, slave=address)
            if response.isError():
                print(f"Error controlling heater ID {address}")
                return False
            return True
        except Exception as e:
            print(f"Error controlling heater ID {address}: {e}")
            return False
        
    
    def update_all_devices(self):
        if getattr(self, "closing", False) or getattr(self, "polling_paused", False):
            return
        """
    Unified method to read all devices efficiently:
    1. Read all Modbus RTU devices (TK4, PSM4) in one batch
    2. Read all ASCII devices (MFC, MFM) via the worker thread
    3. Update all GUI elements
    """
    # === 1. READ ALL MODBUS RTU DEVICES ===

    # Read all TK4 temperature controllers (IDs 1-6)
        temperatures = {}
        for addr in [1, 2, 3, 4, 5, 6]:
            try:
                if self.device_status.get(f"TK4_{addr}", True):
                    temp = self.device_manager.read_temperature(addr)
                    temperatures[addr] = temp
                else:
                    temperatures[addr] = None
            except Exception as e:
                temperatures[addr] = None

    # Read all pressure sensors (PSM4)
        pressures = ["NC"] * 4
        try:
            if self.device_status.get("PSM4", True):
                pressures = self.device_manager.read_pressures()
        except Exception as e:
            print(f"Pressure read error: {e}")
            pressures = ["NC"] * 4

    # === 2. READ ALL ASCII DEVICES (MFC/MFM) VIA WORKER THREAD ===

    # Read MFC flows (ASCII)
        mfc_flows = [None] * 4
        try:
            if self.device_status.get("MFC", True):
                reply_queue = queue.Queue()
                self.button_command_queue.put({"cmd": "read_all_mfc", "reply_queue": reply_queue})
                result = reply_queue.get(timeout=2)
                mfc_flows = result.get("values", [None]*4)
        except Exception as e:
            print(f"MFC read error: {e}")
            mfc_flows = [None] * 4

    # Read MFM flow (ASCII)
        mfm_flow = None
        try:
            if self.device_status.get("MFM", True):
                reply_queue = queue.Queue()
                self.button_command_queue.put({"cmd": "read_mfm", "reply_queue": reply_queue})
                result = reply_queue.get(timeout=2)
                mfm_flow = result.get("value", None)
        except Exception as e:
            print(f"MFM read error: {e}")
            mfm_flow = None

    # Read Power Meter (ASCII, if connected)
        power, energy = None, None

        try:
            if self.device_status.get("PowerMeter", True):
                reply_queue = queue.Queue()
                self.button_command_queue.put({"cmd": "read_power_meter", "reply_queue": reply_queue})
                result = reply_queue.get(timeout=2)
                power = result.get("power", None)
                energy = result.get("energy", None)
        except Exception as e:
            print(f"Power meter read error: {e}")

        scaling = self.settings.get("power_meter_scaling_factor", 1.0)
        if power is not None:
            power *= scaling
        if energy is not None:
            energy *= scaling


    # --- Gas Analyzer ---
        gas_values = None
        try:
            if self.device_status.get("GasAnalyzer", True):
                reply_queue = queue.Queue()
                self.button_command_queue.put({"cmd": "read_gas_analyzer", "reply_queue": reply_queue})
                result = reply_queue.get(timeout=2)
                gas_values = result.get("values", None)
        except Exception as e:
            print(f"Gas analyzer read error: {e}")
            gas_values = None

        data = {
                1: temperatures.get(1),   # Heater (coil, TK4 ID=1)
                2: temperatures.get(2),   # Preheater (TK4 ID=2)
                3: temperatures.get(3),   # Reactor (TK4 ID=3)
                4: temperatures.get(4),   # Temp1 (TK4 ID=4)
                5: temperatures.get(5),   # Temp2 (TK4 ID=5)
                6: temperatures.get(6),   # Temp3 (TK4 ID=6)
                'pressures': pressures[:3],  # Use only first 3 sensors
                'power': power,
                'energy': energy,
                'mfm_flow': mfm_flow,
                'mfc_flows': [mfc_flows[0], mfc_flows[1], mfc_flows[2], mfc_flows[3]] if mfc_flows else [None]*4
                }           

    # === 3. UPDATE ALL GUI ELEMENTS ===

    # Update Heater 1 (ID=2)
        heater_1_temp = temperatures.get(2)
        if heater_1_temp is not None:
            self.settings['heater_1']['pv'] = f"{heater_1_temp:.1f}"
            self.pv_label.SetLabel(f"PV: {heater_1_temp:.1f}")
            self.pv_label.SetForegroundColour(COLOR_RED)
        else:
            self.settings['heater_1']['pv'] = "NC"
            self.pv_label.SetLabel("PV: NC")
            self.pv_label.SetForegroundColour(COLOR_RED)

    # Update Heater 2 (ID=1)
        heater_2_temp = temperatures.get(1)
        if heater_2_temp is not None:
            self.settings["heater_2"]["coil_pv"] = f"{heater_2_temp:.1f}"
            self.coil_pv_label.SetLabel(f"PV: {heater_2_temp:.1f}")
            self.coil_pv_label.SetForegroundColour(COLOR_RED)
        else:
            self.settings["heater_2"]["coil_pv"] = "NC"
            self.coil_pv_label.SetLabel("PV: NC")
            self.coil_pv_label.SetForegroundColour(COLOR_RED)

    # Update Reactor Temp (ID=3)
        reactor_temp = temperatures.get(3)
        if reactor_temp is not None:
            self.settings["heater_2"]["reactor_temp"] = f"{reactor_temp:.1f}"
            self.reactor_temp_label.SetLabel(f"Reactor: {reactor_temp:.1f}")
            self.reactor_temp_label.SetForegroundColour(COLOR_RED)
        else:
            self.settings["heater_2"]["reactor_temp"] = "NC"
            self.reactor_temp_label.SetLabel("Reactor: NC")
            self.reactor_temp_label.SetForegroundColour(COLOR_RED)

    # Update Read-only sensors (IDs 4, 5, 6)
        for i, sensor_id in enumerate([4, 5, 6]):
            sensor_temp = temperatures.get(sensor_id)
            try:
                if (sensor_temp is None or
                    (isinstance(sensor_temp, str) and sensor_temp.upper() == "NC") or
                    (isinstance(sensor_temp, float) and (sensor_temp != sensor_temp))):  # NaN check
                    label_text = "Temp: NC"
                    self.sensor_labels[i].SetForegroundColour(COLOR_RED)
                else:
                    label_text = f"Temp: {float(sensor_temp):.1f}"
                    self.sensor_labels[i].SetForegroundColour(COLOR_RED)
            except Exception:
                label_text = "Temp: NC"
                self.sensor_labels[i].SetForegroundColour(COLOR_RED)
            self.sensor_labels[i].SetLabel(label_text)

    # Update Pressure sensors
        if len(pressures) >= 3:
            for i, pressure_label in enumerate([self.pressure_label, self.pressure_2_label, self.pressure_3_label]):
                if pressures[i] != "NC" and pressures[i] is not None:
                    pressure_label.SetLabel(f"{pressures[i]} bar")
                    pressure_label.SetForegroundColour(COLOR_RED)
                else:
                    pressure_label.SetLabel("-- bar")
                    pressure_label.SetForegroundColour(COLOR_RED)

    # Update MFC flows
        if mfc_flows:
            for i, channel in enumerate(self.mfc_channels):
                if i < len(mfc_flows) and mfc_flows[i] is not None:
                    self.mfc_pv_labels[channel].SetLabel(f"{mfc_flows[i]:.2f}")
                    self.mfc_pv_labels[channel].SetForegroundColour(COLOR_RED)
                else:
                    self.mfc_pv_labels[channel].SetLabel("--")
                    self.mfc_pv_labels[channel].SetForegroundColour(COLOR_RED)

    # Update MFM flow
        if mfm_flow is not None:
            self.mfm_label.SetLabel(f"{mfm_flow:.2f}")
            self.mfm_label.SetForegroundColour(COLOR_RED)
        else:
            self.mfm_label.SetLabel("--")
            self.mfm_label.SetForegroundColour(COLOR_RED)

    # Update Power Meter (if available)
        if power is not None:
            self.power_label.SetLabel(f"{power:.1f}")
            self.power_label.SetForegroundColour(COLOR_RED)
        else:
            self.power_label.SetLabel("--")
            self.power_label.SetForegroundColour(COLOR_RED)

        if energy is not None:
            self.energy_label.SetLabel(f"{energy:.2f}")
            self.energy_label.SetForegroundColour(COLOR_RED)
        else:
            self.energy_label.SetLabel("--")
            self.energy_label.SetForegroundColour(COLOR_RED)

        if gas_values:
            for gas, label in self.gas_labels.items():
                val = gas_values.get(gas)
                if val is not None:
                    label.SetLabel(f"{val:.2f} %")
                    label.SetForegroundColour(COLOR_BLUE)
                else:
                    label.SetLabel("-- %")
                    label.SetForegroundColour(COLOR_RED)
        else:
            for label in self.gas_labels.values():
                label.SetLabel("-- %")
                label.SetForegroundColour(COLOR_RED)

        try:
            self.logger.log(data, gas_values)
        except Exception as e:
            print(f"Logger error: {e}")
    # Schedule next update
        wx.CallLater(2000, self.update_all_devices)   




    def start_device_updates(self):
        self.update_all_devices()


    def update_datetime(self):
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.datetime_label.SetLabel(now)
        wx.CallLater(500, self.update_datetime)

    def start_datetime_updates(self):
        self.update_datetime()

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r") as f:
                settings = json.load(f)
        else:
            settings = {}
        if "heater_1" not in settings:
            settings["heater_1"] = {"sv": "--", "max_temp": "--", "pv": "--"}
        else:
            settings["heater_1"].setdefault("sv", "--")
            settings["heater_1"].setdefault("max_temp", "--")
            settings["heater_1"].setdefault("pv", "--")
        
        return settings

    


    def save_settings(self):
        with open(SETTINGS_FILE, "w") as f:
            json.dump(self.settings, f, indent=4)

    def on_close(self, event):
        self.shutdown_all_devices()
        self.closing = True
        self.running = False
        self.button_command_queue.put({"cmd": "relay_close_all"})
        self.save_settings()
        self.Destroy()
        if hasattr(self, 'worker_thread'):
            print("Waiting for worker thread to exit...")
            self.worker_thread.join(timeout=2)
            print("Worker thread exited.")





    def on_power_meter_click(self, event):
        self.polling_paused = True
        try:
            dlg = PowerMeterDialog(self, self.power_label, self.energy_label, self.integration_status_label)
            dlg.CentreOnParent()
            dlg.ShowModal()
            dlg.Destroy()
        except Exception as e:
            wx.MessageBox(f"Could not open Power Meter dialog: {e}", "Dialog Error", wx.ICON_ERROR)

        self.power_meter_panel.Bind(wx.EVT_LEFT_DOWN, self.on_power_meter_click)


    def on_mouse_click(self, event):
        x, y = event.GetPosition()
        #print(f"Mouse clicked at: {x}, {y} (relative to background)")
        #print(f"Mouse clicked at: {x}, {y} (relative to background)")
        #print(f"Parent window valid: {self.IsShown()}")  # Debugging statement

        # Power Meter area
        if 350 <= x <= 650 and 620 <= y <= 750:
            print("Opening Power Meter Dialog...")
            self.on_power_meter_click(event)
            return

        # Emergency button area
        if 1000 <= x <= 1200 and 150 <= y <= 280:
            print("Emergency button pressed!")
            self.handle_emergency_button()
            return

        
        # Heater 1 area (example coordinates)
        if 310 <= x <= 520 and 60 <= y <= 140:
            try:
                def apply_heater_changes(sv, max_temp, onoff):
                    self.settings["heater_1"]["sv"] = sv
                    self.settings["heater_1"]["max_temp"] = max_temp
                    self.sv_label.SetLabel(f"SV: {sv}" if sv else "SV: --")
                    self.heater_status_label.SetLabel("ON" if onoff else "OFF")
                    self.heater_status_label.SetForegroundColour(COLOR_BLUE if onoff else COLOR_RED)
                self.polling_paused = True
                try:    
                    dialog = HeaterDialog(self, apply_heater_changes, self.settings)
                    dialog.CentreOnParent()
                    dialog.ShowModal()
                    dialog.Destroy()
                finally:
                    self.polling_paused = False
                    self.update_all_devices()
            except Exception as e:
                wx.MessageBox(f"Could not open dialog: {e}", "Dialog Error", wx.ICON_ERROR)
        #
        # Pressure sensor 1 area (example coordinates)
        if 260 <= x <= 300 and 80 <= y <= 120:
            print("Opening Pressure Sensor 1 Dialog...")
            self.open_pressure_dialog(1)

        # Pressure sensor 2 area (example coordinates)
        if 715 <= x <= 760 and 115 <= y <= 140:
            print("Opening Pressure Sensor 2 Dialog...")
            self.open_pressure_dialog(2)

        # Pressure sensor 3 area (example coordinates)
        if 737 <= x <= 770 and 605 <= y <= 630:
            print("Opening Pressure Sensor 3 Dialog...")
            self.open_pressure_dialog(3)

        # Heater 2 area (example coordinates)
        if 660 <= x <= 800 and 175 <= y <= 380:
            print("Opening Heater 2 Dialog...")
            self.polling_paused = True
            try:
                dialog = Heater2Dialog(self, self.coil_sv_label, self.coil_pv_label,
                              self.reactor_temp_label, self.heater_2_status_label, self.settings)
                dialog.CentreOnParent()
                dialog.ShowModal()
                dialog.Destroy()
            except Exception as e:
                wx.MessageBox(f"Could not open dialog: {e}", "Dialog Error", wx.ICON_ERROR)
            finally:
                self.polling_paused = False
                self.update_all_devices()  # Optionally trigger immediate refresh
            gc.collect()

        

        # MFC table area (adjusted coordinates)
        if 100 <= x <= 400 and 240 <= y <= 450:
            print("Opening MFC Settings Dialog...")
            self.open_mfc_settings(event)
        
    def open_pressure_dialog(self, sensor_id):
        self.polling_paused = True
        try:
            dialog = PressureDialog(self, sensor_id, self.settings)
            dialog.CentreOnParent()
            dialog.ShowModal()
            dialog.Destroy()
        except Exception as e:
            wx.MessageBox(f"Could not open Pressure Sensor {sensor_id} dialog: {e}", "Dialog Error", wx.ICON_ERROR)
    
    


    def start_alarm(self):
        try:
            winsound.PlaySound("mixkit-emergency-alert-alarm-1007.wav", winsound.SND_FILENAME | winsound.SND_ASYNC | winsound.SND_LOOP)
        except Exception as e:
            print(f"Alarm sound error: {e}")

    def stop_alarm(self):
        try:       
            winsound.PlaySound(None, winsound.SND_PURGE)
        except Exception as e:
            print(f"Alarm stop error: {e}")
    def stop_alarm(self):
        try:        
            winsound.PlaySound(None, winsound.SND_PURGE)
        except Exception as e:
            print(f"Alarm stop error: {e}") 
        
    def check_emergency_conditions(self):
        if getattr(self, "closing", False):
            return
        try:
        # --- Heater 1 overtemperature (ID=2) ---
            heater_1_temp = self.device_manager.read_temperature(2)
            heater_1_max = self.settings.get("heater_1", {}).get("max_temp", None)
            if (
                heater_1_temp is not None
                and isinstance(heater_1_temp, (int, float))
                and heater_1_max not in (None, "--", "NC", "")
            ):
                try:
                    heater_1_max_val = float(heater_1_max)
                    if heater_1_temp > heater_1_max_val:
                        if not self.overtemp_latched[2]:
                            msg = f"Overtemperature: Heater 1 ({heater_1_temp} > {heater_1_max_val})"
                            print(msg)
                            log_abnormal_event(msg)
                            self.overtemp_latched[2] = True
                            self.handle_emergency_button()
                            return
                    else:
                        self.overtemp_latched[2] = False
                except ValueError:
                    self.overtemp_latched[2] = False
            else:
            # Sensor not connected or invalid: clear latch, no alarm
                self.overtemp_latched[2] = False

        # --- Heater 2 coil overtemperature (ID=1) ---
            heater_2_temp = self.device_manager.read_temperature(1)
            coil_max = self.settings.get("heater_2", {}).get("coil_max_temp", None)
            if (
                heater_2_temp is not None
                and isinstance(heater_2_temp, (int, float))
                and coil_max not in (None, "--", "NC", "")
            ):
                try:
                    coil_max_val = float(coil_max)
                    if heater_2_temp > coil_max_val:
                        if not self.overtemp_latched[1]:
                            msg = f"Overtemperature: Heater 2 ({heater_2_temp} > {coil_max_val})"
                            print(msg)
                            log_abnormal_event(msg)
                            self.overtemp_latched[1] = True
                            self.handle_emergency_button()
                            return
                    else:
                        self.overtemp_latched[1] = False
                except ValueError:
                    self.overtemp_latched[1] = False
            else:
                self.overtemp_latched[1] = False

        # --- Reactor overtemperature (ID=3) ---
            reactor_temp = self.device_manager.read_temperature(3)
            reactor_max = self.settings.get("heater_2", {}).get("reactor_max_temp", None)
            if (
                reactor_temp is not None
                and isinstance(reactor_temp, (int, float))
                and reactor_max not in (None, "--", "NC", "")
            ):
                try:
                    reactor_max_val = float(reactor_max)
                    if reactor_temp > reactor_max_val:
                        if not self.overtemp_latched[3]:
                            msg = f"Overtemperature: Reactor ({reactor_temp} > {reactor_max_val})"
                            print(msg)
                            log_abnormal_event(msg)
                            self.overtemp_latched[3] = True
                            self.handle_emergency_button()
                            return
                    else:
                        self.overtemp_latched[3] = False
                except ValueError:
                    self.overtemp_latched[3] = False
            else:
                self.overtemp_latched[3] = False

        # --- Pressure sensors (IDs 1,2,3) ---
            pressures = self.device_manager.read_pressures()
            sensor_settings = self.settings.get("sensor_settings", {})
            for i, sensor_id in enumerate([1, 2, 3]):
                max_press = sensor_settings.get(f"sensor_{sensor_id}_max_pressure", None)
                value = None
                if (
                    max_press not in (None, "--", "NC", "")
                    and pressures is not None
                    and i < len(pressures)
                    and pressures[i] not in ("NC", "--", None)
                ):
                    try:
                        value = float(pressures[i])
                        max_press_val = float(max_press)
                        if value > max_press_val:
                            if not self.overpress_latched[sensor_id]:
                                msg = f"Overpressure: Sensor {sensor_id} ({value} > {max_press_val})"
                                print(msg)
                                log_abnormal_event(msg)
                                self.overpress_latched[sensor_id] = True
                                self.handle_emergency_button()
                                return
                        else:
                            self.overpress_latched[sensor_id] = False
                    except (ValueError, TypeError):
                        self.overpress_latched[sensor_id] = False
                else:
                # Sensor not connected or invalid: clear latch, no alarm
                    self.overpress_latched[sensor_id] = False

        except Exception as e:
            print(f"Error in check_emergency_conditions: {e}")

        
        wx.CallLater(500, self.check_emergency_conditions)






    def handle_emergency_button(self):
    # Log every Emergency button press (both activation and reset)
        log_abnormal_event("Emergency button pressed")

        if not self.alarm_active:
        # --- EMERGENCY ACTIVATION ---
            print("Emergency button pressed!")
            self.status_bar_label.SetLabel("EMERGENCY STOP: All heaters and flows OFF!")
            self.status_bar_label.SetForegroundColour(COLOR_RED)
            self.alarm_active = True

        # --- PAUSE POLLING IMMEDIATELY ---
            self.polling_paused = True

        # 1. Set all GUI statuses to OFF
            def update_gui_emergency_status():
                self.heater_status_label.SetLabel("OFF")
                self.heater_status_label.SetForegroundColour(COLOR_RED)
                self.heater_2_status_label.SetLabel("OFF")
                self.heater_2_status_label.SetForegroundColour(COLOR_RED)
                for channel in self.mfc_channels:
                    self.mfc_status_labels[channel].SetLabel("OFF")
                    self.mfc_status_labels[channel].SetForegroundColour(COLOR_RED)
            wx.CallAfter(update_gui_emergency_status)

        # 2. Open all relays (ON)
            self.button_command_queue.put({"cmd": "relay_open_all"})

        # 3. Schedule closing all relays after 1 second (OFF)
            def delayed_close():
                time.sleep(1.0)
                self.button_command_queue.put({"cmd": "relay_close_all"})
            threading.Thread(target=delayed_close, daemon=True).start()

        # 4. Stop both heaters (set SV=0, stop command)
            self.button_command_queue.put({"cmd": "set_tk4_sv", "address": 2, "value": 0.0})
            self.button_command_queue.put({"cmd": "stop_tk4", "address": 2})
            self.button_command_queue.put({"cmd": "set_tk4_sv", "address": 1, "value": 0.0})
            self.button_command_queue.put({"cmd": "stop_tk4", "address": 1})

        # 5. Stop all MFC flows
            for i in range(1, 5):  # MFC channels 1-4
                self.button_command_queue.put({"cmd": "set_mfc_flow", "channel": i, "value": 0.0})
                self.button_command_queue.put({"cmd": "on_off_mfc", "channel": i, "state": False})

        # 6. Play alarm sound (loop until reset)
            try:
                winsound.PlaySound(
                    "mixkit-emergency-alert-alarm-1007.wav", 
                    winsound.SND_FILENAME | winsound.SND_ASYNC | winsound.SND_LOOP
                )
            except Exception as e:
                print(f"Alarm sound error: {e}")

        else:
        # --- EMERGENCY RESET ---
            print("Emergency alarm stopped.")
            self.alarm_active = False

        # --- RESUME POLLING ---
            self.polling_paused = False
            wx.CallAfter(self.update_all_devices)

            try:
                winsound.PlaySound(None, winsound.SND_PURGE)
            except Exception as e:
                print(f"Alarm stop error: {e}")

            self.status_bar_label.SetLabel("System Ready")
            self.status_bar_label.SetForegroundColour(COLOR_BLUE)

        # Reset latches
            for k in self.overtemp_latched:
                self.overtemp_latched[k] = False
            for k in self.overpress_latched:
                self.overpress_latched[k] = False

        # Reset GUI statuses just in case
            def update_gui_normal_status():
                self.heater_status_label.SetLabel("OFF")
                self.heater_status_label.SetForegroundColour(COLOR_RED)
                self.heater_2_status_label.SetLabel("OFF")
                self.heater_2_status_label.SetForegroundColour(COLOR_RED)
                for channel in self.mfc_channels:
                    self.mfc_status_labels[channel].SetLabel("OFF")
                    self.mfc_status_labels[channel].SetForegroundColour(COLOR_RED)
            wx.CallAfter(update_gui_normal_status)

        # Resume emergency monitoring
            wx.CallLater(100, self.check_emergency_conditions)

    def open_mfc_settings(self, event):
        self.polling_paused = True
        try:
            dialog = MFCSettingsDialog(self, self.mfc_sv_labels, self.mfc_status_labels, self.settings)
            dialog.CentreOnParent()
            dialog.ShowModal()
            dialog.Destroy()
        except Exception as e:
            wx.MessageBox(f"Could not open MFC Settings dialog: {e}", "Dialog Error", wx.ICON_ERROR)

    def shutdown_all_devices(self):
    # 1. Set all heater SVs to 0 and stop
        for addr in [1, 2]:  # Adjust for your heater IDs
            self.button_command_queue.put({"cmd": "set_tk4_sv", "address": addr, "value": 0.0})
            self.button_command_queue.put({"cmd": "stop_tk4", "address": addr})

    # 2. Set all MFC flows to 0 and turn OFF
        for i in range(1, 5):  # MFC channels 1-4
            self.button_command_queue.put({"cmd": "set_mfc_flow", "channel": i, "value": 0.0})
            self.button_command_queue.put({"cmd": "on_off_mfc", "channel": i, "state": False})

    # 3. Open all relays immediately
        self.button_command_queue.put({"cmd": "relay_open_all"})

    # 4. Schedule closing all relays after 1 second
        def delayed_close():
            time.sleep(1.0)
            self.button_command_queue.put({"cmd": "relay_close_all"})
        threading.Thread(target=delayed_close, daemon=True).start()

            
# --- Main ---
if __name__ == "__main__":
    app = wx.App()
    # === SHARED STYLES FOR FONT AND COLOR (Avoids GDI leaks) ===
    global SHARED_FONT, COLOR_BLUE, COLOR_RED, COLOR_GREEN
    LARGE_FONT = wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
    SHARED_FONT = wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
    COLOR_BLUE = wx.Colour(0, 0, 255)
    COLOR_RED = wx.Colour(255, 0, 0)
    COLOR_GREEN = wx.Colour(0, 255, 0)
    frame = ControlGUI(None, title="POX Control System")
    frame.Show()
    app.MainLoop()
